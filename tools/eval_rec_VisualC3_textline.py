import io
import os
import sys
import torch
import numpy as np
from collections import OrderedDict
from rapidfuzz.distance import Levenshtein
from tqdm import tqdm
from PIL import Image

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

from tools.data import build_dataloader
from tools.engine.config import Config
from tools.engine.trainer import Trainer
from tools.utility import ArgsParser
from tools.utils.ids_syntax import validate_ids_prefix, DEFAULT_IDC_ARITY


# =========================
# 错字检测指标（对齐 + 精简统计）
# =========================
def align_by_opcodes(gt: str, pred: str, gap_char=None):
    """
    将 gt/pred 做全局对齐，返回列对齐序列列表，每项为 (gt_char/gap, pred_char/gap, gt_idx or None)。
    - gt_idx 为该列对应的原始 gt 索引；若该列是 pred 插入（gt 缺失）则为 None。
    """
    aligned = []
    for tag, i1, i2, j1, j2 in Levenshtein.opcodes(gt, pred):
        if tag in ("equal", "replace"):
            len_a = i2 - i1
            len_b = j2 - j1
            m = min(len_a, len_b)
            for k in range(m):
                gi = i1 + k
                pj = j1 + k
                aligned.append((gt[gi], pred[pj], gi))
            if len_a > m:
                for gi in range(i1 + m, i2):
                    aligned.append((gt[gi], gap_char, gi))
            if len_b > m:
                for pj in range(j1 + m, j2):
                    aligned.append((gap_char, pred[pj], None))
        elif tag == "delete":
            for gi in range(i1, i2):
                aligned.append((gt[gi], gap_char, gi))
        elif tag == "insert":
            for pj in range(j1, j2):
                aligned.append((gap_char, pred[pj], None))
        else:
            raise ValueError(f"Unknown opcode tag: {tag}")
    return aligned


def _safe_div(num, den):
    return None if den == 0 else (num / den)


def _safe_pct(num, den):
    v = _safe_div(num, den)
    return None if v is None else (v * 100.0)


def calculate_cuo_metric_compact(gt_sentences, pred_sentences, X='X'):
    """
    仅评 X（错字位点）的检测，输出精简但高信息量的指标：
    - N_clean_sent: GT 无 X 的句子数
    - N_error_sent: GT 有 X 的句子数
    - Char_P / Char_R / Char_F1:
        precision = TP/(TP+FP), recall = TP/(TP+FN)
    - Sent_FA: clean sentence 上预测出任意 X 的比例（误报率）
    - Sent_EM: error sentence 上 exact match（X位置集合完全一致且不能多报/插入X）的比例
    """
    n = min(len(gt_sentences), len(pred_sentences))
    if n == 0:
        return {}

    n_clean_sent = 0
    n_error_sent = 0
    sent_fa = 0   # false alarm on clean sentences
    sent_em = 0   # exact match on error sentences

    # char-level TP/FP/FN（只针对 X）
    tp = 0
    fp = 0
    fn = 0

    for gt, pred in zip(gt_sentences[:n], pred_sentences[:n]):
        aligned = align_by_opcodes(gt, pred, gap_char=None)

        gt_x_pos = set()
        pred_x_pos = set()
        ins_x_cnt = 0

        for gch, pch, gi in aligned:
            g_is_x = (gch == X)
            p_is_x = (pch == X)

            # 统计位置集合（以 GT 坐标为准）
            if gi is not None and g_is_x:
                gt_x_pos.add(gi)

            if p_is_x:
                if gi is None:
                    # pred 在 GT gap（插入列）上输出 X -> 一定是多报(FP) + exact match 必失败
                    ins_x_cnt += 1
                else:
                    pred_x_pos.add(gi)

            # char-level 计数（仅 X 类）
            if gi is None:
                # 插入列：只要插入的是 X，就算 FP
                if p_is_x:
                    fp += 1
            else:
                if g_is_x and p_is_x:
                    tp += 1
                elif (not g_is_x) and p_is_x:
                    fp += 1
                elif g_is_x and (not p_is_x):
                    fn += 1

        gt_has_x = (len(gt_x_pos) > 0)
        pred_has_x = (len(pred_x_pos) > 0) or (ins_x_cnt > 0)

        if gt_has_x:
            n_error_sent += 1
            # exact match：X位置集合完全一致，且不能多报（含插入列X）
            if pred_has_x and (ins_x_cnt == 0) and (pred_x_pos == gt_x_pos):
                sent_em += 1
        else:
            n_clean_sent += 1
            # clean sentence 上只要预测出任意 X 就算误报
            if pred_has_x:
                sent_fa += 1

    # char-level metrics
    char_p = _safe_pct(tp, tp + fp)
    char_r = _safe_pct(tp, tp + fn)
    char_f1 = None
    if char_p is not None and char_r is not None and (char_p + char_r) > 0:
        char_f1 = 2 * char_p * char_r / (char_p + char_r)

    # sentence-level metrics
    sent_fa_rate = _safe_pct(sent_fa, n_clean_sent)  # clean 误报率
    sent_em_rate = _safe_pct(sent_em, n_error_sent)  # error exact-match 率

    return {
        "N_sent": n,
        "N_clean_sent": n_clean_sent,
        "N_error_sent": n_error_sent,
        "Char_P": char_p,
        "Char_R": char_r,
        "Char_F1": char_f1,
        "Sent_FA": sent_fa_rate,
        "Sent_EM": sent_em_rate,
        # 下面三项只为方便定位（可不打印）
        "Char_TP": tp,
        "Char_FP": fp,
        "Char_FN": fn,
    }


def maybe_ids_tokens(tokens):
    """简单启发：若 token 中含常见 IDC 符号，则认为是 IDS 输出。否则视为普通文本，不做合法性统计。"""
    if not tokens:
        return False
    idc_chars = {'⿰', '⿱', '⿲', '⿳', '⿴', '⿵', '⿶', '⿷', '⿸', '⿹', '⿺', '⿻'}
    return any(any(ch in idc_chars for ch in tok) for tok in tokens)


def replace_punctuation(text: str) -> str:
    """将常见中文标点替换为英文标点，保持与参考脚本一致。"""
    if text is None:
        return ''
    mapping = {
        r'，': r',',
        r'。': r'.',
        r'！': r'!',
        r'？': r'?',
        r'；': r';',
        r'：': r':',
        r'“': r'"',
        r'”': r'"',
        r'‘': r"'",
        r'’': r"'",
    }
    for k, v in mapping.items():
        text = text.replace(k, v)
    return text


def to_png_bytes(img_array):
    """将 numpy 图像数组转成 PNG bytes，兼容单通道/三通道及 0~1/0~255 输入。"""
    if img_array is None:
        return None
    arr = np.array(img_array)
    if arr.ndim == 3 and arr.shape[0] in (1, 3):
        arr = np.transpose(arr, (1, 2, 0))
    arr = np.squeeze(arr)
    if arr.max() <= 1.0 + 1e-3:
        arr = arr * 255.0
    arr = np.clip(arr, 0, 255).astype(np.uint8)
    mode = 'L' if arr.ndim == 2 else 'RGB'
    img = Image.fromarray(arr, mode=mode)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    buf.close()
    return data


def split_src_tgt(gt_raw: str):
    """解析 src???tgt 形式标签，缺失时做容错。"""
    if gt_raw is None:
        return '', ''
    parts = str(gt_raw).split('???')
    if len(parts) == 2:
        return parts[0], parts[1]
    return str(gt_raw), ''


def parse_args():
    parser = ArgsParser()
    parser.add_argument(
        '--save_pred_xlsx',
        type=str,
        default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump_textline.xlsx',
    )
    args = parser.parse_args()
    return args


def prepare_cfg(cfg):
    if cfg.cfg['Global']['output_dir'][-1] == '/':
        cfg.cfg['Global']['output_dir'] = cfg.cfg['Global']['output_dir'][:-1]
    if cfg.cfg['Global']['pretrained_model'] is None:
        cfg.cfg['Global']['pretrained_model'] = cfg.cfg['Global']['output_dir'] + '/best.pth'
    cfg.cfg['Global']['use_amp'] = False
    cfg.cfg['PostProcess']['with_ratio'] = True
    cfg.cfg['Metric']['with_ratio'] = True
    cfg.cfg['Metric']['max_len'] = 30
    cfg.cfg['Global']['max_text_length'] = 30
    cfg.cfg['Metric']['max_ratio'] = 12
    keep_keys = cfg.cfg['Eval']['dataset']['transforms'][-1]['KeepKeys']['keep_keys']
    if 'real_ratio' not in keep_keys:
        keep_keys.append('real_ratio')
    return cfg


def dump_predictions(trainer, datadir, output_log, dataset_name, image_bytes, det_inputs=None, idc_arity=None):
    config_each = trainer.cfg.copy()
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    model.eval()
    num = 0
    src_true_num = 0
    tgt_true_num = 0
    src_ned_list = []
    tgt_ned_list = []
    if idc_arity is None:
        idc_arity = DEFAULT_IDC_ARITY
    legal_token = 0
    total_token = 0
    legal_seq = 0
    total_seq = 0
    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        for batch_idx, batch in enumerate(valid_dataloader):
            batch_tensor = [t.to(device) for t in batch[:3]]
            batch_numpy = [t.numpy() for t in batch[:3]]
            raw_images = batch[3] if len(batch) > 3 else None
            preds = model(batch_tensor[0], data=batch_tensor[1:])
            post_result = post_process(preds, batch_numpy)
            texts, gts = post_result if isinstance(post_result, tuple) else (post_result, None)
            # print(f"texts:{texts}, gts:{gts}")  # debug
            for i, (txt, prob) in enumerate(texts):
                gt_text = ''
                if gts is not None and i < len(gts):
                    gt_text = gts[i][0]
                gt_src_txt, gt_tgt_txt = split_src_tgt(gt_text)

                txt_norm = replace_punctuation(txt)
                gt_src_norm = replace_punctuation(gt_src_txt)
                gt_tgt_norm = replace_punctuation(gt_tgt_txt)
                src_ned = 1 - Levenshtein.normalized_distance(txt_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                tgt_ned = 1 - Levenshtein.normalized_distance(txt_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_list.append(src_ned)
                tgt_ned_list.append(tgt_ned)
                num += 1
                if int(src_ned) == 1:
                    src_true_num += 1
                if int(tgt_ned) == 1:
                    tgt_true_num += 1

                # IDS 合法性统计（若模型输出为 IDS 序列，则按空格切分 token 逐一校验）
                tokens = [tok for tok in str(txt).strip().split() if tok]
                if maybe_ids_tokens(tokens):
                    seq_ok = True
                    for tok in tokens:
                        ok, need, _ = validate_ids_prefix([ch for ch in tok if not ch.isspace()], idc_arity=idc_arity, require_closed=True)
                        total_token += 1
                        if ok:
                            legal_token += 1
                        else:
                            seq_ok = False
                    if tokens:
                        total_seq += 1
                        if seq_ok:
                            legal_seq += 1

                # 收集错字检测输入（默认以 src 标签中的 X 作为错位标记）
                if det_inputs is not None:
                    det_inputs['gts'].append(gt_src_norm)
                    det_inputs['preds'].append(txt_norm)

                img_name = f"{dataset_name}_{sample_offset + i}"
                output_log['img_name'].append(img_name)
                output_log['type'].append(dataset_name)
                output_log['label_src'].append(gt_src_norm)
                output_log['label_tgt'].append(gt_tgt_norm)
                output_log['pred'].append(txt_norm)
                output_log['NED_src'].append(float(src_ned))
                output_log['NED_tgt'].append(float(tgt_ned))
                try:
                    sample_img = raw_images[i] if raw_images is not None else None
                    image_bytes.append(to_png_bytes(sample_img))
                except Exception:
                    image_bytes.append(None)
            sample_offset += len(texts)
            pbar.update(1)
        pbar.close()
    model.train()
    src_pnacc = src_true_num / num if num else 0.0
    tgt_pnacc = tgt_true_num / num if num else 0.0
    src_ned_mean = float(np.mean(src_ned_list)) if src_ned_list else 0.0
    tgt_ned_mean = float(np.mean(tgt_ned_list)) if tgt_ned_list else 0.0
    legality = {
        'legal_token': legal_token,
        'total_token': total_token,
        'legal_seq': legal_seq,
        'total_seq': total_seq,
    }
    return src_pnacc, tgt_pnacc, src_ned_mean, tgt_ned_mean, num, legality


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)
    FLAGS = vars(FLAGS)
    opt = FLAGS.pop('opt')
    cfg.merge_dict(FLAGS)
    cfg.merge_dict(opt)
    cfg = prepare_cfg(cfg)

    save_pred_xlsx = FLAGS.get('save_pred_xlsx')
    if save_pred_xlsx is None:
        save_pred_xlsx = os.path.join(cfg.cfg['Global']['output_dir'], 'preds_dump_textline.xlsx')
    os.makedirs(os.path.dirname(save_pred_xlsx), exist_ok=True)

    trainer = Trainer(cfg, mode='eval')

    data_dirs_list = []
    if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
        data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
    else:
        data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
        if data_dir_single:
            data_dirs_list = [[data_dir_single]]

    # 默认同时评测 textline test_correct 与 test_faked
    data_dirs_list = [[
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_correct',
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_faked',
    ]]

    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label_src', []),
        ('label_tgt', []),
        ('pred', []),
        ('NED_src', []),
        ('NED_tgt', []),
    ])
    det_inputs = {'gts': [], 'preds': []}
    image_bytes = []
    every_src_PNacc_list = []
    every_tgt_PNacc_list = []
    every_src_ned_list = []
    every_tgt_ned_list = []
    total_num = 0
    total_src_True_num = 0
    total_tgt_True_num = 0
    total_src_ned_list = []
    total_tgt_ned_list = []
    legal_token_sum = 0
    total_token_sum = 0
    legal_seq_sum = 0
    total_seq_sum = 0
    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            src_pnacc, tgt_pnacc, src_ned_mean, tgt_ned_mean, num, legality = dump_predictions(trainer, datadir, output_log, dataset_name, image_bytes, det_inputs=det_inputs, idc_arity=DEFAULT_IDC_ARITY)
            print(f"{dataset_name}:\t\t src_acc: {100 * src_pnacc:6g}, src_norm_edit_dis:{100 * src_ned_mean:6g}, tgt_acc: {100 * tgt_pnacc:6g}, tgt_norm_edit_dis:{100 * tgt_ned_mean:6g}")
            every_src_PNacc_list.append(src_pnacc)
            every_tgt_PNacc_list.append(tgt_pnacc)
            every_src_ned_list.append(src_ned_mean)
            every_tgt_ned_list.append(tgt_ned_mean)
            total_num += num
            total_src_True_num += int(src_pnacc * num)
            total_src_ned_list.extend([src_ned_mean] * num)
            total_tgt_True_num += int(tgt_pnacc * num)
            total_tgt_ned_list.extend([tgt_ned_mean] * num)
            legal_token_sum += legality['legal_token']
            total_token_sum += legality['total_token']
            legal_seq_sum += legality['legal_seq']
            total_seq_sum += legality['total_seq']

    try:
        import pandas as pd
        df = pd.DataFrame(output_log)
        df.to_excel(save_pred_xlsx, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.utils import get_column_letter

            wb = load_workbook(save_pred_xlsx)
            ws = wb.active
            img_col = ws.max_column + 1
            ws.cell(row=1, column=img_col, value='image')
            img_col_letter = get_column_letter(img_col)
            target_px = 160
            ws.column_dimensions[img_col_letter].width = max(ws.column_dimensions[img_col_letter].width or 0, target_px / 7.0)
            embedded = 0
            for r_idx, data in enumerate(image_bytes, start=2):
                if not data:
                    continue
                img_obj = OpenpyxlImage(io.BytesIO(data))
                try:
                    with Image.open(io.BytesIO(data)) as pil_img:
                        w, h = pil_img.size
                except Exception:
                    w, h = None, None
                img_obj.width = target_px
                if w and h and w > 0:
                    img_obj.height = h * (target_px / float(w))
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, img_obj.height * 0.75)
                img_obj.anchor = f"{img_col_letter}{r_idx}"
                ws.add_image(img_obj)
                embedded += 1
            print(f"[INFO] Embedded {embedded} images into Excel")
            wb.save(save_pred_xlsx)
        except Exception as embed_err:
            print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")

        total_src_acc = (total_src_True_num / total_num) if total_num else 0.0
        total_src_ned = float(np.mean(total_src_ned_list)) if total_src_ned_list else 0.0
        s_mean_src_acc = float(np.mean(every_src_PNacc_list)) if every_src_PNacc_list else 0.0
        s_mean_src_ned = float(np.mean(every_src_ned_list)) if every_src_ned_list else 0.0
        total_tgt_acc = (total_tgt_True_num / total_num) if total_num else 0.0
        total_tgt_ned = float(np.mean(total_tgt_ned_list)) if total_tgt_ned_list else 0.0
        s_mean_tgt_acc = float(np.mean(every_tgt_PNacc_list)) if every_tgt_PNacc_list else 0.0
        s_mean_tgt_ned = float(np.mean(every_tgt_ned_list)) if every_tgt_ned_list else 0.0
        s_weight_src_acc = float(np.sum(np.array(every_src_PNacc_list))) if every_src_PNacc_list else 0.0
        s_weight_src_ned = float(np.sum(np.array(every_src_ned_list))) if every_src_ned_list else 0.0
        s_weight_tgt_acc = float(np.sum(np.array(every_tgt_PNacc_list))) if every_tgt_PNacc_list else 0.0
        s_weight_tgt_ned = float(np.sum(np.array(every_tgt_ned_list))) if every_tgt_ned_list else 0.0

        print(f"total:\t\t src_acc: {100 * total_src_acc:6g}, src_norm_edit_dis:{100 * total_src_ned:6g},tgt_acc: {100 * total_tgt_acc:6g}, tgt_norm_edit_dis:{100 * total_tgt_ned:6g}")
        print(f"S_mean:\t\t src_acc: {100 * s_mean_src_acc:6g}, src_norm_edit_dis:{100 * s_mean_src_ned:6g}, tgt_acc: {100 * s_mean_tgt_acc:6g}, tgt_norm_edit_dis:{100 * s_mean_tgt_ned:6g}")
        print(f"S_weight:\t\t src_acc: {100 * s_weight_src_acc:6g}, src_norm_edit_dis:{100 * s_weight_src_ned:6g}, tgt_acc: {100 * s_weight_tgt_acc:6g}, tgt_norm_edit_dis:{100 * s_weight_tgt_ned:6g}")
        print(f'Predictions (with NED) saved to {save_pred_xlsx}')
        if total_token_sum:
            token_legal_rate = legal_token_sum / total_token_sum
            print(f"IDS token legality: {token_legal_rate * 100:.2f}% ({legal_token_sum}/{total_token_sum})")
        if total_seq_sum:
            seq_legal_rate = legal_seq_sum / total_seq_sum
            print(f"IDS sequence legality: {seq_legal_rate * 100:.2f}% ({legal_seq_sum}/{total_seq_sum})")
        if (total_token_sum == 0) and (total_seq_sum == 0):
            print("[INFO] IDS legality skipped (no IDC tokens detected in predictions).")
    except Exception as e:
        print(f'[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.')

    # ========= 精简的错字检测指标（使用标签中的 X） =========
    det = calculate_cuo_metric_compact(det_inputs['gts'], det_inputs['preds'], X='X')

    def fmt(x):
        return "N/A" if x is None else f"{x:.3f}"

    print("\nCuo detection metrics (compact, mixed clean+error):")
    if det:
        print(f"N_sent={det['N_sent']} | clean={det['N_clean_sent']} | error={det['N_error_sent']}")
        print(f"Char_P={fmt(det['Char_P'])}%  Char_R={fmt(det['Char_R'])}%  Char_F1={fmt(det['Char_F1'])}%")
        print(f"Sent_FA={fmt(det['Sent_FA'])}%  Sent_EM={fmt(det['Sent_EM'])}%")
        # 如需排查，可临时打开这一行
        # print(f"(debug) Char_TP={det['Char_TP']} Char_FP={det['Char_FP']} Char_FN={det['Char_FN']}")
    else:
        print("No sentences to evaluate cuo detection.")


if __name__ == '__main__':
    main()
