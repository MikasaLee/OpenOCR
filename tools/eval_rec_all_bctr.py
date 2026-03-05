import io
import os
import sys
import torch
import numpy as np
from collections import OrderedDict
from rapidfuzz.distance import Levenshtein
from tqdm import tqdm
from PIL import Image

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

from tools.data import build_dataloader  
from tools.engine.config import Config  
from tools.engine.trainer import Trainer  
from tools.utility import ArgsParser  

S_WEIGHT = np.array([0.25,0.25,0.25,0.25], dtype=np.float32)


def split_tensor_prefix(batch, device):
    """Use contiguous tensor prefix as model/post-process inputs."""
    tensor_end = 0
    for item in batch:
        if torch.is_tensor(item):
            tensor_end += 1
        else:
            break
    if tensor_end == 0:
        raise RuntimeError('No tensor fields found in batch. Check KeepKeys/collate_fn.')
    batch_tensor = [t.to(device) for t in batch[:tensor_end]]
    batch_numpy = [t.detach().cpu().numpy() for t in batch[:tensor_end]]
    extra_fields = batch[tensor_end:]
    return batch_tensor, batch_numpy, extra_fields


def get_keep_keys_from_cfg(config_each):
    transforms = config_each.get('Eval', {}).get('dataset', {}).get(
        'transforms', [])
    for t in transforms:
        if 'KeepKeys' in t:
            return t['KeepKeys'].get('keep_keys', [])
    return []


def filter_meta_tensor_keys(batch_tensor, batch_numpy, keep_keys):
    """Remove tensor meta keys that should not go into model/post-process."""
    if not keep_keys:
        return batch_tensor, batch_numpy
    meta_keys = {'real_ratio'}
    valid_len = min(len(batch_tensor), len(keep_keys))
    drop_idx = [
        i for i, k in enumerate(keep_keys[:valid_len]) if k in meta_keys
    ]
    if not drop_idx:
        return batch_tensor, batch_numpy
    keep_idx = [i for i in range(len(batch_tensor)) if i not in set(drop_idx)]
    batch_tensor = [batch_tensor[i] for i in keep_idx]
    batch_numpy = [batch_numpy[i] for i in keep_idx]
    return batch_tensor, batch_numpy

def normalize_text(text) -> str:
    """Normalize decoded text payload to plain string."""
    if text is None:
        return ''
    if isinstance(text, str):
        return text
    if isinstance(text, bytes):
        return text.decode('utf-8', errors='ignore')
    if isinstance(text, (list, tuple)):
        parts = []
        for item in text:
            if isinstance(item, bytes):
                parts.append(item.decode('utf-8', errors='ignore'))
            elif isinstance(item, (list, tuple)):
                parts.append(normalize_text(item))
            else:
                parts.append(str(item))
        return ''.join(parts)
    return str(text)


def to_float(v, default=1.0):
    try:
        return float(v)
    except Exception:
        return default


def normalize_text_pairs(items):
    """Normalize predictions to [(text, prob), ...]."""
    if items is None:
        return []
    if not isinstance(items, (list, tuple)):
        return [(normalize_text(items), 1.0)]

    out = []
    for it in items:
        if isinstance(it, (list, tuple)) and len(it) >= 1:
            # Typical format: (text, prob)
            if len(it) >= 2 and not isinstance(it[1], (list, tuple, dict)):
                out.append((normalize_text(it[0]), to_float(it[1], 1.0)))
            else:
                out.append((normalize_text(it[0]), 1.0))
        else:
            out.append((normalize_text(it), 1.0))
    return out


def normalize_gt_list(items):
    """Normalize labels to [text, ...]."""
    if items is None:
        return None
    if not isinstance(items, (list, tuple)):
        return [normalize_text(items)]

    out = []
    for it in items:
        if isinstance(it, (list, tuple)) and len(it) >= 1:
            out.append(normalize_text(it[0]))
        else:
            out.append(normalize_text(it))
    return out


def parse_post_result(post_result, use_gtc_decode=False, infer_branch='gtc'):
    """Unify post-process output to (texts, gts)."""
    selected = post_result

    # GTCLabelDecode: [gtc_result, ctc_result]
    if use_gtc_decode and isinstance(post_result, list) and len(post_result) == 2:
        selected = post_result[0] if infer_branch == 'gtc' else post_result[1]

    # CharWiseVerifyPostProcess with batch:
    # [(text_preds, text_labels), (ids_preds, ids_labels)]
    if (isinstance(selected, list) and len(selected) == 2 and
            all(isinstance(x, (list, tuple)) and len(x) == 2 for x in selected)):
        selected = selected[0]

    # Dict output (some custom post-process inference modes)
    if isinstance(selected, dict):
        texts = selected.get('text', [])
        gts = selected.get('label_text', None)
        if gts is None:
            gts = selected.get('label', None)
        return normalize_text_pairs(texts), normalize_gt_list(gts)

    # Common output: (texts, gts)
    if (isinstance(selected, (list, tuple)) and len(selected) == 2 and
            isinstance(selected[0], (list, tuple))):
        return normalize_text_pairs(selected[0]), normalize_gt_list(selected[1])

    # Fallback: texts only
    return normalize_text_pairs(selected), None


def replace_punctuation(text) -> str:
    """将常见中文标点替换为英文标点，保持与参考脚本一致。"""
    text = normalize_text(text)
    if not text:
        return ''
    # 中文到英文的简单映射
    mapping = {
        r'，': r',', 
        r'。': r'.',
        r'！': r'!', 
        r'？': r'?', 
        r'；': r';', 
        r'：': r':',
        r'“': r'"', 
        r'”': r'"', 
        r'‘': r"'", 
        r'’': r"'",
    }
    for k, v in mapping.items():
        text = text.replace(k, v)
    return text


def to_png_bytes(img_array):
    """将 numpy 图像数组转成 PNG bytes，兼容单通道/三通道及 0~1/0~255 输入。"""
    if img_array is None:
        return None

    arr = np.array(img_array)
    if arr.ndim == 3 and arr.shape[0] in (1, 3):
        arr = np.transpose(arr, (1, 2, 0))
    arr = np.squeeze(arr)

    if arr.max() <= 1.0 + 1e-3:
        arr = arr * 255.0
    arr = np.clip(arr, 0, 255).astype(np.uint8)

    mode = 'L' if arr.ndim == 2 else 'RGB'
    img = Image.fromarray(arr, mode=mode)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    buf.close()
    return data



def parse_args():
    parser = ArgsParser()
    parser.add_argument(
        '--infer_branch',
        type=str,
        default='ctc',
        choices=['gtc', 'ctc'],
        help='Inference branch for GTCDecoder.',
    )
    parser.add_argument(
        '--save_pred_xlsx',
        type=str,
        default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump.xlsx',
    )
    parser.add_argument(
        '--no_save_xlsx',
        action='store_true',
        help='Disable XLSX export (fastest mode).',
    )
    parser.add_argument(
        '--no_embed_image',
        action='store_true',
        help='Export XLSX without embedded images (faster).',
    )
    args = parser.parse_args()
    return args


def prepare_cfg(cfg, infer_branch='gtc', enable_raw_image=True):
    # Align with eval_rec_all_ch tweaks
    if cfg.cfg['Global']['output_dir'][-1] == '/':
        cfg.cfg['Global']['output_dir'] = cfg.cfg['Global']['output_dir'][:-1]
    if cfg.cfg['Global']['pretrained_model'] is None:
        cfg.cfg['Global']['pretrained_model'] = cfg.cfg['Global']['output_dir'] + '/best.pth'
    cfg.cfg['Global']['use_amp'] = False
    # cfg.cfg['PostProcess']['with_ratio'] = True
    # cfg.cfg['Metric']['with_ratio'] = True
    cfg.cfg['Metric']['max_len'] = 25
    cfg.cfg['Metric']['max_ratio'] = 12

    eval_transforms = cfg.cfg['Eval']['dataset']['transforms']

    # 增强 KeepKeys
    keep_keys_list = None
    for t in eval_transforms:
        if 'KeepKeys' in t:
            keep_keys_list = t['KeepKeys']['keep_keys']
            break
    if keep_keys_list is None and len(eval_transforms) > 0 and 'KeepKeys' in eval_transforms[-1]:
        keep_keys_list = eval_transforms[-1]['KeepKeys']['keep_keys']
    if keep_keys_list is not None:
        if 'real_ratio' not in keep_keys_list:
            keep_keys_list.append('real_ratio')

    # ---------------------------------------------------------
    # 按需注入 Eval 保存原图所需配置 (SaveRawImageBytes, KeepKeys, collate_fn)
    # ---------------------------------------------------------
    if enable_raw_image:
        if not any('SaveRawImageBytes' in t for t in eval_transforms):
            insert_idx = 0
            for i, t in enumerate(eval_transforms):
                if 'DecodeImagePIL' in t:
                    insert_idx = i + 1
                    break
            eval_transforms.insert(
                insert_idx,
                {'SaveRawImageBytes': {
                    'dst_key': 'image_raw',
                    'src_key': 'image'
                }})

        if 'loader' not in cfg.cfg['Eval']:
            cfg.cfg['Eval']['loader'] = {}
        cfg.cfg['Eval']['loader']['collate_fn'] = 'RecWithRawCollator'

        if keep_keys_list is not None and 'image_raw' not in keep_keys_list:
            keep_keys_list.append('image_raw')

    decoder_name = cfg.cfg.get('Architecture', {}).get('Decoder', {}).get('name')
    post_name = cfg.cfg.get('PostProcess', {}).get('name')
    if decoder_name == 'GTCDecoder' and post_name == 'GTCLabelDecode':
        # GTC eval requires dual-branch outputs from model.
        cfg.cfg['Architecture']['Decoder']['infer_gtc'] = True
        cfg.cfg['PostProcess']['only_gtc'] = False
    # 不引入非张量键，避免 to(device) 失败；img_name 以索引生成
    return cfg


def dump_predictions(trainer,
                     datadir,
                     output_log,
                     dataset_name,
                     image_bytes,
                     infer_branch='gtc',
                     collect_raw_images=True):
    config_each = trainer.cfg.copy()
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    use_gtc_decode = trainer.cfg.get('PostProcess', {}).get('name') == 'GTCLabelDecode'
    keep_keys = get_keep_keys_from_cfg(config_each)
    model.eval()
    num = 0
    true_num = 0
    ned_list = []
    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        for batch_idx, batch in enumerate(valid_dataloader):
            batch_tensor, batch_numpy, extra_fields = split_tensor_prefix(
                batch, device)
            batch_tensor, batch_numpy = filter_meta_tensor_keys(
                batch_tensor, batch_numpy, keep_keys)

            raw_images = None
            if collect_raw_images:
                for item in extra_fields:
                    if isinstance(item, list) and len(item) > 0:
                        first_sample = item[0]
                        if isinstance(first_sample, np.ndarray) and first_sample.ndim >= 2:
                            raw_images = item
                            break
                        if isinstance(first_sample, bytes):
                            raw_images = item
                            break

            preds = model(batch_tensor[0], data=batch_tensor[1:])
            post_result = post_process(preds, batch_numpy)
            texts, gts = parse_post_result(post_result,
                                           use_gtc_decode=use_gtc_decode,
                                           infer_branch=infer_branch)

            for i, (txt, prob) in enumerate(texts):
                gt_text = ''
                if gts is not None and i < len(gts):
                    gt_text = gts[i]
                # 标点标准化后再计算 NED
                txt_norm = replace_punctuation(txt)
                gt_norm = replace_punctuation(gt_text)
                ned = 1 - Levenshtein.normalized_distance(txt_norm, gt_norm) if gt_norm is not None else 0.0
                ned_list.append(ned)
                num += 1
                if int(ned) == 1:
                    true_num += 1
                # 输出格式对齐：img_name, type, label, pred, NED
                img_name = f"{dataset_name}_{sample_offset + i}"
                output_log['img_name'].append(img_name)
                output_log['type'].append(dataset_name)
                output_log['label'].append(gt_norm)
                output_log['pred'].append(txt_norm)
                output_log['NED'].append(float(ned))
                if collect_raw_images and image_bytes is not None:
                    try:
                        sample_img = raw_images[i] if raw_images is not None else None
                        image_bytes.append(to_png_bytes(sample_img))
                    except Exception:
                        image_bytes.append(None)
            sample_offset += len(texts)
            pbar.update(1)
        pbar.close()
    model.train()
    pnacc = true_num / num if num else 0.0
    ned_mean = float(np.mean(ned_list)) if ned_list else 0.0
    return pnacc, ned_mean, num


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)
    FLAGS = vars(FLAGS)
    infer_branch = FLAGS.get('infer_branch', 'gtc')
    no_save_xlsx = FLAGS.get('no_save_xlsx', False)
    no_embed_image = FLAGS.get('no_embed_image', False)
    if no_save_xlsx:
        no_embed_image = True

    opt = FLAGS.pop('opt')
    cfg.merge_dict(FLAGS)
    cfg.merge_dict(opt)
    cfg = prepare_cfg(cfg,
                      infer_branch=infer_branch,
                      enable_raw_image=(not no_embed_image))

    save_pred_xlsx = None
    if not no_save_xlsx:
        save_pred_xlsx = FLAGS.get('save_pred_xlsx')
        if save_pred_xlsx is None:
            save_pred_xlsx = os.path.join(cfg.cfg['Global']['output_dir'], 'preds_dump.xlsx')
        save_dir = os.path.dirname(save_pred_xlsx)
        if save_dir:
            os.makedirs(save_dir, exist_ok=True)

    trainer = Trainer(cfg, mode='eval')
    trainer.logger.info(f'Inference branch: {infer_branch}')
    trainer.logger.info(
        f'XLSX export: {"off" if no_save_xlsx else "on"}, embed image: {"off" if no_embed_image else "on"}')
    if trainer.cfg.get('PostProcess', {}).get('name') != 'GTCLabelDecode':
        trainer.logger.info(
            f'Inference branch={infer_branch} ignored (PostProcess={trainer.cfg.get("PostProcess", {}).get("name")}).')

    data_dirs_list = []
    if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
        data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
    else:
        data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
        if data_dir_single:
            data_dirs_list = [[data_dir_single]]

    data_dirs_list = [[
      r'/a800data1/lirunrui/origin_datasets/bchw_dataset/scene/scene_test',
      r'/a800data1/lirunrui/origin_datasets/bchw_dataset/web/web_test',
      r'/a800data1/lirunrui/origin_datasets/bchw_dataset/document/document_test',
      r'/a800data1/lirunrui/origin_datasets/bchw_dataset/hw/hw_test',
    ]]
    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label', []),
        ('pred', []),
        ('NED', []),
    ])
    image_bytes = [] if (not no_save_xlsx and not no_embed_image) else None
    every_PNacc_list = []
    every_ned_list = []
    total_num = 0
    total_True_num = 0
    total_ned_list = []
    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            pnacc, ned_mean, num = dump_predictions(
                trainer,
                datadir,
                output_log,
                dataset_name,
                image_bytes,
                infer_branch=infer_branch,
                collect_raw_images=(not no_save_xlsx and not no_embed_image),
            )
            print(f"{dataset_name}:\t\t acc: {100 * pnacc:6g}, norm_edit_dis:{100 * ned_mean:6g}")
            every_PNacc_list.append(pnacc)
            every_ned_list.append(ned_mean)
            total_num += num
            total_True_num += int(pnacc * num)
            total_ned_list.extend([ned_mean] * num)

    try:
        if no_save_xlsx:
            total_acc = (total_True_num / total_num) if total_num else 0.0
            total_ned = float(np.mean(total_ned_list)) if total_ned_list else 0.0
            s_mean_acc = float(np.mean(every_PNacc_list)) if every_PNacc_list else 0.0
            s_mean_ned = float(np.mean(every_ned_list)) if every_ned_list else 0.0
            if every_PNacc_list:
                acc_arr = np.array(every_PNacc_list, dtype=np.float32)
                ned_arr = np.array(every_ned_list, dtype=np.float32)
                if len(acc_arr) == len(S_WEIGHT):
                    weights = S_WEIGHT / np.sum(S_WEIGHT)
                    s_weight_acc = float(np.sum(acc_arr * weights))
                    s_weight_ned = float(np.sum(ned_arr * weights))
                else:
                    print(f"[WARN] S_WEIGHT length mismatch: metrics={len(acc_arr)}, weights={len(S_WEIGHT)}. Fallback to S_mean.")
                    s_weight_acc = s_mean_acc
                    s_weight_ned = s_mean_ned
            else:
                s_weight_acc = 0.0
                s_weight_ned = 0.0
            print(f"total:\t\t acc: {100 * total_acc:6g}, norm_edit_dis:{100 * total_ned:6g}")
            print(f"S_mean:\t\t acc: {100 * s_mean_acc:6g}, norm_edit_dis:{100 * s_mean_ned:6g}")
            print(f"S_weight:\t\t acc: {100 * s_weight_acc:6g}, norm_edit_dis:{100 * s_weight_ned:6g}")
            print('[INFO] --no_save_xlsx enabled, skip XLSX export.')
            return

        import pandas as pd
        df = pd.DataFrame(output_log)
        df.to_excel(save_pred_xlsx, index=False)
        try:
            if not no_embed_image:
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image as OpenpyxlImage
                from openpyxl.utils import get_column_letter

                wb = load_workbook(save_pred_xlsx)
                ws = wb.active
                img_col = ws.max_column + 1
                ws.cell(row=1, column=img_col, value='image')
                img_col_letter = get_column_letter(img_col)

                target_px = 160
                ws.column_dimensions[img_col_letter].width = max(ws.column_dimensions[img_col_letter].width or 0, target_px / 7.0)

                embedded = 0
                for r_idx, data in enumerate(image_bytes, start=2):
                    if not data:
                        continue
                    img_obj = OpenpyxlImage(io.BytesIO(data))
                    try:
                        with Image.open(io.BytesIO(data)) as pil_img:
                            w, h = pil_img.size
                    except Exception:
                        w, h = None, None

                    img_obj.width = target_px
                    if w and h and w > 0:
                        img_obj.height = h * (target_px / float(w))
                    ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, img_obj.height * 0.75)

                    img_obj.anchor = f"{img_col_letter}{r_idx}"
                    ws.add_image(img_obj)
                    embedded += 1

                print(f"[INFO] Embedded {embedded} images into Excel")
                wb.save(save_pred_xlsx)
            else:
                print('[INFO] --no_embed_image enabled, exported XLSX without images.')
        except Exception as embed_err:
            print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")
        # 汇总日志
        total_acc = (total_True_num / total_num) if total_num else 0.0
        total_ned = float(np.mean(total_ned_list)) if total_ned_list else 0.0
        s_mean_acc = float(np.mean(every_PNacc_list)) if every_PNacc_list else 0.0
        s_mean_ned = float(np.mean(every_ned_list)) if every_ned_list else 0.0
        if every_PNacc_list:
            acc_arr = np.array(every_PNacc_list, dtype=np.float32)
            ned_arr = np.array(every_ned_list, dtype=np.float32)
            if len(acc_arr) == len(S_WEIGHT):
                weights = S_WEIGHT / np.sum(S_WEIGHT)
                s_weight_acc = float(np.sum(acc_arr * weights))
                s_weight_ned = float(np.sum(ned_arr * weights))
            else:
                print(f"[WARN] S_WEIGHT length mismatch: metrics={len(acc_arr)}, weights={len(S_WEIGHT)}. Fallback to S_mean.")
                s_weight_acc = s_mean_acc
                s_weight_ned = s_mean_ned
        else:
            s_weight_acc = 0.0
            s_weight_ned = 0.0
        print(f"total:\t\t acc: {100 * total_acc:6g}, norm_edit_dis:{100 * total_ned:6g}")
        print(f"S_mean:\t\t acc: {100 * s_mean_acc:6g}, norm_edit_dis:{100 * s_mean_ned:6g}")
        print(f"S_weight:\t\t acc: {100 * s_weight_acc:6g}, norm_edit_dis:{100 * s_weight_ned:6g}")
        print(f'Predictions (with NED) saved to {save_pred_xlsx}')
    except Exception as e:
        print(f'[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.')


if __name__ == '__main__':
    main()
