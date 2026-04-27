import io
import os
import sys
import math
import torch
import torch.nn.functional as F
import numpy as np
from collections import OrderedDict
from rapidfuzz.distance import Levenshtein
from tqdm import tqdm
from PIL import Image

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

import openrec.preprocess as preprocess_module
from tools.data import build_dataloader
from tools.engine.config import Config
from tools.engine.trainer import Trainer
from tools.utility import ArgsParser


CONF_THR = 0.2


class SaveLabelText:

    def __init__(self, dst_key='label_raw', src_key='label', **kwargs):
        self.dst_key = dst_key
        self.src_key = src_key

    def __call__(self, data):
        data[self.dst_key] = data.get(self.src_key)
        return data


preprocess_module.SaveLabelText = SaveLabelText


# ============== Helpers ==============

def replace_punctuation(text: str) -> str:
    if text is None:
        return ''
    mapping = {
        '，': ',', '。': '.', '！': '!', '？': '?', '；': ';', '：': ':',
        '“': '"', '”': '"', '‘': "'", '’': "'",
    }
    for k, v in mapping.items():
        text = text.replace(k, v)
    return text


def to_png_bytes(img_array):
    if img_array is None:
        return None
    arr = np.array(img_array)
    if arr.ndim == 3 and arr.shape[0] in (1, 3):
        arr = np.transpose(arr, (1, 2, 0))
    arr = np.squeeze(arr)
    if arr.max() <= 1.0 + 1e-3:
        arr = arr * 255.0
    arr = np.clip(arr, 0, 255).astype(np.uint8)
    mode = 'L' if arr.ndim == 2 else 'RGB'
    img = Image.fromarray(arr, mode=mode)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    buf.close()
    return data


def split_src_tgt(gt_raw: str):
    if gt_raw is None:
        return '', ''
    # print("gt_raw:", gt_raw)
    parts = str(gt_raw).split('\t')  # 应该是\t，但ids字典中没有就转成了<unk>
    # print("parts:", parts)
    if len(parts) == 2:
        return parts[0], parts[1]
    return str(gt_raw), ''


def _safe_div(num, den):
    return None if den == 0 else (num / den)


def _safe_pct(num, den):
    v = _safe_div(num, den)
    return None if v is None else (v * 100.0)


def align_by_opcodes(gt: str, pred: str, gap_char=None):
    aligned = []
    for tag, i1, i2, j1, j2 in Levenshtein.opcodes(gt, pred):
        if tag in ('equal', 'replace'):
            len_a = i2 - i1
            len_b = j2 - j1
            m = min(len_a, len_b)
            for k in range(m):
                gi = i1 + k
                pj = j1 + k
                aligned.append((gt[gi], pred[pj], gi))
            if len_a > m:
                for gi in range(i1 + m, i2):
                    aligned.append((gt[gi], gap_char, gi))
            if len_b > m:
                for pj in range(j1 + m, j2):
                    aligned.append((gap_char, pred[pj], None))
        elif tag == 'delete':
            for gi in range(i1, i2):
                aligned.append((gt[gi], gap_char, gi))
        elif tag == 'insert':
            for pj in range(j1, j2):
                aligned.append((gap_char, pred[pj], None))
        else:
            raise ValueError(f'Unknown opcode tag: {tag}')
    return aligned


def calculate_cuo_metric_compact(gt_sentences, pred_sentences, X='X'):
    n = min(len(gt_sentences), len(pred_sentences))
    if n == 0:
        return {}

    n_clean_sent = 0
    n_error_sent = 0
    sent_fa = 0
    sent_em = 0

    tp = 0
    fp_align = 0
    fp_ins = 0
    fn = 0
    tn = 0

    for gt, pred in zip(gt_sentences[:n], pred_sentences[:n]):
        aligned = align_by_opcodes(gt, pred, gap_char=None)

        gt_x_pos = set()
        pred_x_pos = set()
        ins_x_cnt = 0

        for gch, pch, gi in aligned:
            g_is_x = (gch == X)
            p_is_x = (pch == X)

            if gi is not None and g_is_x:
                gt_x_pos.add(gi)

            if p_is_x:
                if gi is None:
                    ins_x_cnt += 1
                else:
                    pred_x_pos.add(gi)

            if gi is None:
                if p_is_x:
                    fp_ins += 1
            else:
                if g_is_x:
                    if p_is_x:
                        tp += 1
                    else:
                        fn += 1
                else:
                    if p_is_x:
                        fp_align += 1
                    else:
                        tn += 1

        gt_has_x = (len(gt_x_pos) > 0)
        pred_has_x = (len(pred_x_pos) > 0) or (ins_x_cnt > 0)

        if gt_has_x:
            n_error_sent += 1
            if pred_has_x and (ins_x_cnt == 0) and (pred_x_pos == gt_x_pos):
                sent_em += 1
        else:
            n_clean_sent += 1
            if pred_has_x:
                sent_fa += 1

    fp_total = fp_align + fp_ins
    char_p = _safe_pct(tp, tp + fp_total)
    char_r = _safe_pct(tp, tp + fn)
    char_f1 = None
    if char_p is not None and char_r is not None and (char_p + char_r) > 0:
        char_f1 = 2 * char_p * char_r / (char_p + char_r)

    spec = _safe_div(tn, tn + fp_align)
    sens = _safe_div(tp, tp + fn)
    char_bal_acc = None
    if spec is not None and sens is not None:
        char_bal_acc = (spec + sens) / 2.0 * 100.0

    char_mcc = None
    numerator = (tp * tn) - (fp_align * fn)
    denominator = (tp + fp_align) * (tp + fn) * (tn + fp_align) * (tn + fn)
    if denominator > 0:
        char_mcc = numerator / math.sqrt(denominator)

    sent_fa_rate = _safe_pct(sent_fa, n_clean_sent)
    sent_em_rate = _safe_pct(sent_em, n_error_sent)

    return {
        'N_sent': n,
        'N_clean_sent': n_clean_sent,
        'N_error_sent': n_error_sent,
        'Char_P': char_p,
        'Char_R': char_r,
        'Char_F1': char_f1,
        'Char_BalAcc': char_bal_acc,
        'Char_MCC': char_mcc,
        'Sent_FA': sent_fa_rate,
        'Sent_EM': sent_em_rate,
    }


def fmt(x):
    return 'N/A' if x is None else f'{x:.3f}'


def split_batch(batch, device):
    tensor_end = 0
    for item in batch:
        if torch.is_tensor(item):
            tensor_end += 1
        else:
            break
    batch_tensor = [t.to(device) for t in batch[:tensor_end]]
    batch_numpy = [t.numpy() for t in batch[:tensor_end]]
    batch_extra = batch[tensor_end:]
    return batch_tensor, batch_numpy, batch_extra


def find_raw_images(extra_fields):
    for item in extra_fields:
        if isinstance(item, list) and len(item) > 0:
            first = item[0]
            if isinstance(first, np.ndarray) and first.ndim >= 2:
                return item
            if isinstance(first, bytes):
                return item
    return None


def find_raw_labels(extra_fields):
    for item in extra_fields:
        if isinstance(item, list) and len(item) > 0:
            first = item[0]
            if isinstance(first, (str, bytes)):
                return item
    return None


def _extract_logits_for_char_conf(preds, decoder_name):
    if isinstance(preds, torch.Tensor):
        return preds

    if decoder_name == 'LISTERLabelDecode':
        if isinstance(preds, (list, tuple)) and len(preds) > 1 and isinstance(
                preds[1], dict):
            logits = preds[1].get('logits', None)
            if torch.is_tensor(logits) or isinstance(logits, np.ndarray):
                return logits

    if isinstance(preds, dict):
        logits = preds.get('logits', None)
        if torch.is_tensor(logits) or isinstance(logits, np.ndarray):
            return logits
        align_logits = preds.get('align', None)
        if isinstance(align_logits, list) and len(align_logits) > 0 and (
                torch.is_tensor(align_logits[-1]) or
                isinstance(align_logits[-1], np.ndarray)):
            return align_logits[-1]
        vision_logits = preds.get('vision', None)
        if torch.is_tensor(vision_logits) or isinstance(vision_logits,
                                                         np.ndarray):
            return vision_logits

    if isinstance(preds, (list, tuple)):
        for item in preds:
            if torch.is_tensor(item) and item.ndim == 3:
                return item
            if isinstance(item, np.ndarray) and item.ndim == 3:
                return item
            if isinstance(item, dict):
                logits = item.get('logits', None)
                if torch.is_tensor(logits) or isinstance(logits, np.ndarray):
                    return logits

    return preds


def _ensure_probs(preds):
    if isinstance(preds, torch.Tensor):
        probs = preds.detach()
    else:
        probs = torch.as_tensor(preds)
    if probs.ndim != 3:
        raise ValueError(f'Expected [B, T, C], got {tuple(probs.shape)}')
    if probs.numel() == 0:
        return probs.float()
    sample = probs[0, 0]
    looks_like_prob = bool(torch.all(sample >= -1e-6)) and bool(torch.all(sample <= 1.0 + 1e-3)) and abs(float(sample.sum().item()) - 1.0) < 1e-2
    if looks_like_prob:
        return probs.float()
    return F.softmax(probs.float(), dim=-1)


def decode_text_with_char_conf(preds, post_process):
    decoder_name = post_process.__class__.__name__
    logits = _extract_logits_for_char_conf(preds, decoder_name)
    probs = _ensure_probs(logits).cpu().numpy()
    pred_idx = probs.argmax(axis=2)
    pred_prob = probs.max(axis=2)

    texts = []
    confs = []

    if decoder_name == 'CTCLabelDecode':
        ignored_tokens = set(post_process.get_ignored_tokens())
        for b in range(pred_idx.shape[0]):
            selection = np.ones(len(pred_idx[b]), dtype=bool)
            if len(selection) > 1:
                selection[1:] = pred_idx[b][1:] != pred_idx[b][:-1]
            for ignored in ignored_tokens:
                selection &= pred_idx[b] != ignored
            chars = [post_process.character[int(idx)] for idx in pred_idx[b][selection]]
            texts.append(''.join(chars))
            confs.append(pred_prob[b][selection].astype(np.float32).tolist())
        return texts, confs

    eos_tokens = {'</s>', '<eos>', 'eos'}
    skip_tokens = {'<s>', '<pad>', 'sos', 'blank', '<INF>', '<INB>'}
    for b in range(pred_idx.shape[0]):
        chars = []
        char_confs = []
        for t in range(pred_idx.shape[1]):
            idx = int(pred_idx[b][t])
            if idx < 0 or idx >= len(post_process.character):
                continue
            token = post_process.character[idx]
            if token in eos_tokens:
                break
            if token in skip_tokens:
                continue
            chars.append(token)
            char_confs.append(float(pred_prob[b][t]))
        texts.append(''.join(chars))
        confs.append(char_confs)
    return texts, confs


def build_confdet_text(text, char_conf, thr=CONF_THR):
    if not text:
        return ''
    if not char_conf:
        return text
    out = []
    for i, ch in enumerate(text):
        conf = char_conf[i] if i < len(char_conf) else 1.0
        out.append('X' if conf < thr else ch)
    return ''.join(out)


def parse_args():
    parser = ArgsParser()
    parser.add_argument('--save_pred_xlsx', type=str, default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump_text_confdet.xlsx')
    args = parser.parse_args()
    return args


def prepare_cfg(cfg):
    if cfg.cfg['Global']['output_dir'][-1] == '/':
        cfg.cfg['Global']['output_dir'] = cfg.cfg['Global']['output_dir'][:-1]
    if cfg.cfg['Global']['pretrained_model'] is None:
        cfg.cfg['Global']['pretrained_model'] = cfg.cfg['Global']['output_dir'] + '/best.pth'
    cfg.cfg['Global']['use_amp'] = False
    cfg.cfg['PostProcess']['with_ratio'] = True
    cfg.cfg['Metric']['with_ratio'] = True
    cfg.cfg['Metric']['max_len'] = 30
    cfg.cfg['Global']['max_text_length'] = 30
    cfg.cfg['Metric']['max_ratio'] = 12

    eval_transforms = cfg.cfg['Eval']['dataset']['transforms']

    if not any('SaveRawImageBytes' in t for t in eval_transforms):
        insert_idx = 0
        for i, t in enumerate(eval_transforms):
            if 'DecodeImagePIL' in t:
                insert_idx = i + 1
                break
        eval_transforms.insert(insert_idx, {'SaveRawImageBytes': {'dst_key': 'image_raw', 'src_key': 'image'}})

    if not any('SaveLabelText' in t for t in eval_transforms):
        insert_idx = len(eval_transforms)
        for i, t in enumerate(eval_transforms):
            op_name = list(t.keys())[0]
            if 'LabelEncode' in op_name:
                insert_idx = i
                break
        eval_transforms.insert(insert_idx, {'SaveLabelText': {'dst_key': 'label_raw', 'src_key': 'label'}})

    if 'loader' not in cfg.cfg['Eval']:
        cfg.cfg['Eval']['loader'] = {}
    cfg.cfg['Eval']['loader']['collate_fn'] = 'RecWithRawCollator'

    keep_keys_list = None
    for t in eval_transforms:
        if 'KeepKeys' in t:
            keep_keys_list = t['KeepKeys']['keep_keys']
            break
    if keep_keys_list is None and len(eval_transforms) > 0 and 'KeepKeys' in eval_transforms[-1]:
        keep_keys_list = eval_transforms[-1]['KeepKeys']['keep_keys']

    if keep_keys_list is not None:
        if 'real_ratio' not in keep_keys_list:
            keep_keys_list.append('real_ratio')
        if 'image_raw' not in keep_keys_list:
            keep_keys_list.append('image_raw')
        if 'label_raw' not in keep_keys_list:
            keep_keys_list.append('label_raw')

    return cfg


# ============== Core Dump ==============

def dump_predictions(trainer, datadir, output_log, dataset_name, image_bytes, det_inputs_text, det_inputs_confdet, conf_thr=CONF_THR):
    config_each = trainer.cfg.copy()
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    model.eval()

    num = 0
    src_true_text = 0
    tgt_true_text = 0
    src_ned_text = []
    tgt_ned_text = []

    src_true_confdet = 0
    tgt_true_confdet = 0
    src_ned_confdet = []
    tgt_ned_confdet = []

    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        for batch_idx, batch in enumerate(valid_dataloader):
            batch_tensor, batch_numpy, batch_extra = split_batch(batch, device)
            raw_images = find_raw_images(batch_extra)
            raw_labels = find_raw_labels(batch_extra)

            model_data = batch_tensor[1:3] if len(batch_tensor) >= 3 else batch_tensor[1:]
            preds = model(batch_tensor[0], data=model_data)
            post_result = post_process(preds, batch_numpy)
            texts = post_result[0] if isinstance(post_result, tuple) else post_result
            # texts = texts[0][0]     # only for DTRNet-only_text
        
            try:
                pred_texts, pred_char_confs = decode_text_with_char_conf(preds, post_process)
            except Exception as exc:
                trainer.logger.warning(f'Char-confidence decode failed on {dataset_name}: {exc}')
                pred_texts = []
                pred_char_confs = []

            def norm_pairs(lst):
                out = []
                for item in lst:
                    if isinstance(item, (tuple, list)) and len(item) >= 1:
                        s = item[0]
                        p = item[1] if len(item) > 1 else 1.0
                        out.append((str(s), float(p)))
                    else:
                        out.append((str(item), 1.0))
                return out

            text_list = norm_pairs(texts)

            for i in range(len(text_list)):
                pred_text_post, _ = text_list[i]
                pred_text = pred_texts[i] if i < len(pred_texts) and pred_texts[i] != '' else pred_text_post
                char_conf = pred_char_confs[i] if i < len(pred_char_confs) else []
                pred_confdet = build_confdet_text(pred_text, char_conf, conf_thr)

                gt_text = raw_labels[i] if raw_labels is not None and i < len(raw_labels) else ''
                if isinstance(gt_text, bytes):
                    gt_text = gt_text.decode('utf-8', errors='ignore')
                gt_src_txt, gt_tgt_txt = split_src_tgt(gt_text)

                pred_text_norm = replace_punctuation(pred_text)
                pred_confdet_norm = replace_punctuation(pred_confdet)
                gt_src_norm = replace_punctuation(gt_src_txt)
                gt_tgt_norm = replace_punctuation(gt_tgt_txt)

                s_ned_text = 1 - Levenshtein.normalized_distance(pred_text_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                t_ned_text = 1 - Levenshtein.normalized_distance(pred_text_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_text.append(s_ned_text)
                tgt_ned_text.append(t_ned_text)
                if int(s_ned_text) == 1:
                    src_true_text += 1
                if int(t_ned_text) == 1:
                    tgt_true_text += 1

                s_ned_confdet = 1 - Levenshtein.normalized_distance(pred_confdet_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                t_ned_confdet = 1 - Levenshtein.normalized_distance(pred_confdet_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_confdet.append(s_ned_confdet)
                tgt_ned_confdet.append(t_ned_confdet)
                if int(s_ned_confdet) == 1:
                    src_true_confdet += 1
                if int(t_ned_confdet) == 1:
                    tgt_true_confdet += 1

                num += 1

                det_inputs_text['gts'].append(gt_src_norm)
                det_inputs_text['preds'].append(pred_text_norm)
                det_inputs_confdet['gts'].append(gt_src_norm)
                det_inputs_confdet['preds'].append(pred_confdet_norm)

                img_name = f"{dataset_name}_{sample_offset + i}"
                output_log['img_name'].append(img_name)
                output_log['type'].append(dataset_name)
                output_log['label_src'].append(gt_src_norm)
                output_log['label_tgt'].append(gt_tgt_norm)
                output_log['pred_text'].append(pred_text_norm)
                output_log['pred_confdet'].append(pred_confdet_norm)
                output_log['char_conf'].append(' '.join(f'{c:.4f}' for c in char_conf))
                output_log['NED_text_src'].append(float(s_ned_text))
                output_log['NED_text_tgt'].append(float(t_ned_text))
                output_log['NED_confdet_src'].append(float(s_ned_confdet))
                output_log['NED_confdet_tgt'].append(float(t_ned_confdet))

                try:
                    sample_img = raw_images[i] if raw_images is not None else None
                    image_bytes.append(to_png_bytes(sample_img))
                except Exception:
                    image_bytes.append(None)

            sample_offset += len(text_list)
            pbar.update(1)
        pbar.close()

    model.train()

    return (
        src_true_text / num if num else 0.0,
        tgt_true_text / num if num else 0.0,
        float(np.mean(src_ned_text)) if src_ned_text else 0.0,
        float(np.mean(tgt_ned_text)) if tgt_ned_text else 0.0,
        src_true_confdet / num if num else 0.0,
        tgt_true_confdet / num if num else 0.0,
        float(np.mean(src_ned_confdet)) if src_ned_confdet else 0.0,
        float(np.mean(tgt_ned_confdet)) if tgt_ned_confdet else 0.0,
        num,
    )


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)
    FLAGS = vars(FLAGS)
    opt = FLAGS.pop('opt')
    cfg.merge_dict(FLAGS)
    cfg.merge_dict(opt)
    cfg = prepare_cfg(cfg)

    save_pred_xlsx = FLAGS.get('save_pred_xlsx')
    if save_pred_xlsx is None:
        save_pred_xlsx = os.path.join(cfg.cfg['Global']['output_dir'], 'preds_dump_text_confdet.xlsx')
    os.makedirs(os.path.dirname(save_pred_xlsx), exist_ok=True)

    trainer = Trainer(cfg, mode='eval')
    trainer.logger.info(f'Confidence threshold fixed to {CONF_THR} (Visual-C3 OCR heuristic)')
    if trainer.cfg.get('PostProcess', {}).get('name') == 'GTCLabelDecode':
        raise NotImplementedError('eval_rec_textline_confdet.py follows eval_rec_textline_triple.py and only supports single-branch postprocess, e.g. CTCLabelDecode/ARLabelDecode.')

    data_dirs_list = []
    if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
        data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
    else:
        data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
        if data_dir_single:
            data_dirs_list = [[data_dir_single]]

    data_dirs_list = [[
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_correct',
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_fakedv2',
    ]]

    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label_src', []),
        ('label_tgt', []),
        ('pred_text', []),
        ('pred_confdet', []),
        ('char_conf', []),
        ('NED_text_src', []),
        ('NED_text_tgt', []),
        ('NED_confdet_src', []),
        ('NED_confdet_tgt', []),
    ])

    det_inputs_text = {'gts': [], 'preds': []}
    det_inputs_confdet = {'gts': [], 'preds': []}
    image_bytes = []

    total_num = 0
    total_src_text_true = 0
    total_tgt_text_true = 0
    total_src_text_neds = []
    total_tgt_text_neds = []
    total_src_confdet_true = 0
    total_tgt_confdet_true = 0
    total_src_confdet_neds = []
    total_tgt_confdet_neds = []

    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            (
                src_pnacc_text, tgt_pnacc_text, src_ned_text_mean, tgt_ned_text_mean,
                src_pnacc_confdet, tgt_pnacc_confdet, src_ned_confdet_mean, tgt_ned_confdet_mean,
                num,
            ) = dump_predictions(
                trainer, datadir, output_log, dataset_name, image_bytes,
                det_inputs_text, det_inputs_confdet, conf_thr=CONF_THR,
            )

            print(f"{dataset_name}:\t text_src_acc: {100*src_pnacc_text:6g}, text_src_NED:{100*src_ned_text_mean:6g}, text_tgt_acc: {100*tgt_pnacc_text:6g}, text_tgt_NED:{100*tgt_ned_text_mean:6g}")
            print(f"{dataset_name}:\t confdet_src_acc: {100*src_pnacc_confdet:6g}, confdet_src_NED:{100*src_ned_confdet_mean:6g}, confdet_tgt_acc: {100*tgt_pnacc_confdet:6g}, confdet_tgt_NED:{100*tgt_ned_confdet_mean:6g}")

            total_num += num
            total_src_text_true += int(src_pnacc_text * num)
            total_tgt_text_true += int(tgt_pnacc_text * num)
            total_src_text_neds.extend([src_ned_text_mean] * num)
            total_tgt_text_neds.extend([tgt_ned_text_mean] * num)
            total_src_confdet_true += int(src_pnacc_confdet * num)
            total_tgt_confdet_true += int(tgt_pnacc_confdet * num)
            total_src_confdet_neds.extend([src_ned_confdet_mean] * num)
            total_tgt_confdet_neds.extend([tgt_ned_confdet_mean] * num)

    try:
        import pandas as pd
        df = pd.DataFrame(output_log)
        df.to_excel(save_pred_xlsx, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.utils import get_column_letter
            wb = load_workbook(save_pred_xlsx)
            ws = wb.active
            img_col = ws.max_column + 1
            ws.cell(row=1, column=img_col, value='image')
            img_col_letter = get_column_letter(img_col)
            target_px = 160
            ws.column_dimensions[img_col_letter].width = max(ws.column_dimensions[img_col_letter].width or 0, target_px / 7.0)
            embedded = 0
            for r_idx, data in enumerate(image_bytes, start=2):
                if not data:
                    continue
                img_obj = OpenpyxlImage(io.BytesIO(data))
                try:
                    with Image.open(io.BytesIO(data)) as pil_img:
                        w, h = pil_img.size
                except Exception:
                    w, h = None, None
                img_obj.width = target_px
                if w and h and w > 0:
                    img_obj.height = h * (target_px / float(w))
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, img_obj.height * 0.75)
                img_obj.anchor = f"{img_col_letter}{r_idx}"
                ws.add_image(img_obj)
                embedded += 1
            print(f"[INFO] Embedded {embedded} images into Excel")
            wb.save(save_pred_xlsx)
        except Exception as embed_err:
            print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")

        total_src_text_acc = (total_src_text_true / total_num) if total_num else 0.0
        total_tgt_text_acc = (total_tgt_text_true / total_num) if total_num else 0.0
        total_src_text_ned = float(np.mean(total_src_text_neds)) if total_src_text_neds else 0.0
        total_tgt_text_ned = float(np.mean(total_tgt_text_neds)) if total_tgt_text_neds else 0.0

        total_src_confdet_acc = (total_src_confdet_true / total_num) if total_num else 0.0
        total_tgt_confdet_acc = (total_tgt_confdet_true / total_num) if total_num else 0.0
        total_src_confdet_ned = float(np.mean(total_src_confdet_neds)) if total_src_confdet_neds else 0.0
        total_tgt_confdet_ned = float(np.mean(total_tgt_confdet_neds)) if total_tgt_confdet_neds else 0.0

        print(f"total (text):\t src_acc: {100*total_src_text_acc:6g}, src_NED:{100*total_src_text_ned:6g}, tgt_acc: {100*total_tgt_text_acc:6g}, tgt_NED:{100*total_tgt_text_ned:6g}")
        print(f"total (confdet):\t src_acc: {100*total_src_confdet_acc:6g}, src_NED:{100*total_src_confdet_ned:6g}, tgt_acc: {100*total_tgt_confdet_acc:6g}, tgt_NED:{100*total_tgt_confdet_ned:6g}")

        det_text = calculate_cuo_metric_compact(det_inputs_text['gts'], det_inputs_text['preds'], X='X')
        det_confdet = calculate_cuo_metric_compact(det_inputs_confdet['gts'], det_inputs_confdet['preds'], X='X')

        print("\nCuo detection metrics (text):")
        if det_text:
            print(f"N_sent={det_text['N_sent']} | clean={det_text['N_clean_sent']} | error={det_text['N_error_sent']}")
            print(f"Char_P={fmt(det_text['Char_P'])}%  Char_R={fmt(det_text['Char_R'])}%  Char_F1={fmt(det_text['Char_F1'])}%")
            print(f"Char_BalAcc={fmt(det_text['Char_BalAcc'])}%  Char_MCC={fmt(det_text['Char_MCC'])}")
            print(f"Sent_FA={fmt(det_text['Sent_FA'])}%  Sent_EM={fmt(det_text['Sent_EM'])}%")
        else:
            print("No sentences for text branch.")

        print("\nCuo detection metrics (confdet):")
        if det_confdet:
            print(f"N_sent={det_confdet['N_sent']} | clean={det_confdet['N_clean_sent']} | error={det_confdet['N_error_sent']}")
            print(f"Char_P={fmt(det_confdet['Char_P'])}%  Char_R={fmt(det_confdet['Char_R'])}%  Char_F1={fmt(det_confdet['Char_F1'])}%")
            print(f"Char_BalAcc={fmt(det_confdet['Char_BalAcc'])}%  Char_MCC={fmt(det_confdet['Char_MCC'])}")
            print(f"Sent_FA={fmt(det_confdet['Sent_FA'])}%  Sent_EM={fmt(det_confdet['Sent_EM'])}%")
        else:
            print("No sentences for confdet branch.")

        print(f"Predictions saved to {save_pred_xlsx}")
    except Exception as e:
        print(f"[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.")


if __name__ == '__main__':
    main()
