import io
import copy
import os
import sys
import time
import torch
import numpy as np
from collections import OrderedDict
from rapidfuzz.distance import Levenshtein
from tqdm import tqdm
from PIL import Image

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

from tools.data import build_dataloader  
from tools.engine.config import Config  
from tools.engine.trainer import Trainer  
from tools.utility import ArgsParser  
from tools.eval_efficiency_utils import (
    finalize_efficiency_stats,
    init_efficiency_stats,
    maybe_profile_first_batch,
    merge_efficiency_stats,
    print_efficiency_summary,
    replace_gtc_decoder_with_gtc_only,
    split_batch_tensors,
    sync_device,
    update_efficiency_stats,
)

S_WEIGHT = np.array([1], dtype=np.float32)

def replace_punctuation(text: str) -> str:
    """Normalize common Chinese punctuation to English punctuation."""
    if text is None:
        return ''
    mapping = {
        '，': ',',
        '。': '.',
        '！': '!',
        '？': '?',
        '；': ';',
        '：': ':',
        '“': '"',
        '”': '"',
        '‘': "'",
        '’': "'",
    }
    for k, v in mapping.items():
        text = text.replace(k, v)
    return text


def sequence_edit_distance(pred_seq, gt_seq):
    """Levenshtein distance for token sequences."""
    if not pred_seq:
        return len(gt_seq)
    if not gt_seq:
        return len(pred_seq)

    prev = list(range(len(gt_seq) + 1))
    for i, pred_token in enumerate(pred_seq, start=1):
        curr = [i]
        for j, gt_token in enumerate(gt_seq, start=1):
            cost = 0 if pred_token == gt_token else 1
            curr.append(min(
                prev[j] + 1,
                curr[j - 1] + 1,
                prev[j - 1] + cost,
            ))
        prev = curr
    return prev[-1]


def init_htr_error_stats():
    return {
        'char_dist': 0,
        'char_count': 0,
        'word_dist': 0,
        'word_count': 0,
    }


def update_htr_error_stats(stats, pred_text, gt_text):
    pred_text = pred_text or ''
    gt_text = gt_text or ''

    char_dist = Levenshtein.distance(pred_text, gt_text)
    char_count = len(gt_text)
    pred_words = pred_text.split()
    gt_words = gt_text.split()
    word_dist = sequence_edit_distance(pred_words, gt_words)
    word_count = len(gt_words)

    stats['char_dist'] += char_dist
    stats['char_count'] += char_count
    stats['word_dist'] += word_dist
    stats['word_count'] += word_count

    sample_cer = char_dist / char_count if char_count else float(char_dist > 0)
    sample_wer = word_dist / word_count if word_count else float(word_dist > 0)
    return sample_cer, sample_wer


def merge_htr_error_stats(total_stats, stats):
    if total_stats is None:
        total_stats = init_htr_error_stats()
    for key in total_stats:
        total_stats[key] += stats.get(key, 0)
    return total_stats


def finalize_htr_error_stats(stats):
    char_dist = stats.get('char_dist', 0)
    char_count = stats.get('char_count', 0)
    word_dist = stats.get('word_dist', 0)
    word_count = stats.get('word_count', 0)
    return {
        'cer': char_dist / char_count if char_count else float(char_dist > 0),
        'wer': word_dist / word_count if word_count else float(word_dist > 0),
    }


def to_png_bytes(img_array):
    """Convert numpy image array to PNG bytes."""
    if img_array is None:
        return None

    arr = np.array(img_array)
    if arr.ndim == 3 and arr.shape[0] in (1, 3):
        arr = np.transpose(arr, (1, 2, 0))
    arr = np.squeeze(arr)

    if arr.max() <= 1.0 + 1e-3:
        arr = arr * 255.0
    arr = np.clip(arr, 0, 255).astype(np.uint8)

    mode = 'L' if arr.ndim == 2 else 'RGB'
    img = Image.fromarray(arr, mode=mode)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    buf.close()
    return data



def parse_args():
    parser = ArgsParser()
    parser.add_argument(
        '--infer_branch',
        type=str,
        default='ctc',
        choices=['ctc', 'gtc'],
        help='Inference branch for GTCDecoder.',
    )
    parser.add_argument(
        '--save_pred_xlsx',
        type=str,
        default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump.xlsx',
    )
    args = parser.parse_args()
    return args


def prepare_cfg(cfg, infer_branch='ctc'):
    # Align with eval_rec_all_ch tweaks
    if cfg.cfg['Global']['output_dir'][-1] == '/':
        cfg.cfg['Global']['output_dir'] = cfg.cfg['Global']['output_dir'][:-1]
    if cfg.cfg['Global']['pretrained_model'] is None:
        cfg.cfg['Global']['pretrained_model'] = cfg.cfg['Global']['output_dir'] + '/best.pth'
    cfg.cfg['Global']['use_amp'] = False
    # cfg.cfg['PostProcess']['with_ratio'] = True
    # cfg.cfg['Metric']['with_ratio'] = True
    cfg.cfg['Metric']['max_len'] = 100
    cfg.cfg['Metric']['max_ratio'] = 12
    keep_keys = cfg.cfg['Eval']['dataset']['transforms'][-1]['KeepKeys']['keep_keys']
    if 'real_ratio' not in keep_keys:
        keep_keys.append('real_ratio')

    decoder_name = cfg.cfg.get('Architecture', {}).get('Decoder', {}).get('name')
    post_name = cfg.cfg.get('PostProcess', {}).get('name')
    if decoder_name == 'GTCDecoder' and infer_branch == 'ctc':
        cfg.cfg['Architecture']['Decoder']['infer_gtc'] = False
        if post_name == 'GTCLabelDecode':
            cfg.cfg['PostProcess'] = {
                'name': 'CTCLabelDecode',
                'character_dict_path': cfg.cfg['Global']['character_dict_path'],
                'use_space_char': cfg.cfg['Global']['use_space_char'],
            }
        if cfg.cfg.get('Metric', {}).get('name') == 'RecGTCMetric':
            cfg.cfg['Metric']['name'] = 'RecMetric'
            cfg.cfg['Metric']['main_indicator'] = 'acc'
    elif infer_branch == 'gtc':
        if decoder_name == 'GTCDecoder':
            cfg.cfg['Architecture']['Decoder']['infer_gtc'] = True
        if post_name == 'GTCLabelDecode':
            cfg.cfg['PostProcess']['only_gtc'] = True
    return cfg


def select_postprocess_batch(post_process, batch_numpy):
    if post_process.__class__.__name__ == 'CTCLabelDecode' and len(batch_numpy) >= 5:
        return [None] + batch_numpy[-2:]
    return batch_numpy


def dump_predictions(trainer, datadir, output_log, dataset_name, image_bytes, infer_branch='ctc'):
    config_each = trainer.cfg.copy()
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    use_gtc_decode = trainer.cfg.get('PostProcess', {}).get('name') == 'GTCLabelDecode'
    model.eval()
    num = 0
    true_num = 0
    ned_list = []
    htr_stats = init_htr_error_stats()
    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        for batch_idx, batch in enumerate(valid_dataloader):
            batch_tensor = [t.to(device) for t in batch[:5]] # 3 for svtrv2 or 5 for eduor
            batch_numpy = [t.numpy() for t in batch[:5]] # 3 or 5
            raw_images = batch[3] if len(batch) > 3 else None
            preds = model(batch_tensor[0], data=batch_tensor[1:])
            post_result = post_process(preds, batch_numpy)

            selected_result = post_result
            if use_gtc_decode and isinstance(post_result, list) and len(post_result) == 2:
                # GTCLabelDecode returns [gtc_result, ctc_result]
                selected_result = post_result[0] if infer_branch == 'gtc' else post_result[1]

            if isinstance(selected_result, tuple):
                texts, gts = selected_result
            else:
                texts, gts = selected_result, None

            for i, (txt, prob) in enumerate(texts):
                gt_text = ''
                if gts is not None and i < len(gts):
                    # gts elements are (text, prob)
                    gt_text = gts[i][0]
                # 鏍囩偣鏍囧噯鍖栧悗鍐嶈绠?NED
                txt_norm = replace_punctuation(txt)
                gt_norm = replace_punctuation(gt_text)
                ned = 1 - Levenshtein.normalized_distance(txt_norm, gt_norm) if gt_norm is not None else 0.0
                cer, wer = update_htr_error_stats(htr_stats, txt_norm, gt_norm)
                ned_list.append(ned)
                num += 1
                if int(ned) == 1:
                    true_num += 1
                # 杈撳嚭鏍煎紡瀵归綈锛歩mg_name, type, label, pred, NED
                img_name = f"{dataset_name}_{sample_offset + i}"
                output_log['img_name'].append(img_name)
                output_log['type'].append(dataset_name)
                output_log['label'].append(gt_norm)
                output_log['pred'].append(txt_norm)
                output_log['NED'].append(float(ned))
                output_log['CER'].append(float(cer))
                output_log['WER'].append(float(wer))
                try:
                    sample_img = raw_images[i] if raw_images is not None else None
                    image_bytes.append(to_png_bytes(sample_img))
                except Exception:
                    image_bytes.append(None)
            sample_offset += len(texts)
            pbar.update(1)
        pbar.close()
    model.train()
    pnacc = true_num / num if num else 0.0
    ned_mean = float(np.mean(ned_list)) if ned_list else 0.0
    return pnacc, ned_mean, num, htr_stats


def dump_predictions_with_efficiency(trainer, datadir, output_log, dataset_name, image_bytes, infer_branch='ctc'):
    config_each = copy.deepcopy(trainer.cfg)
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    use_gtc_decode = trainer.cfg.get('PostProcess', {}).get('name') == 'GTCLabelDecode'
    model.eval()
    num = 0
    true_num = 0
    ned_list = []
    htr_stats = init_htr_error_stats()
    eff_stats = init_efficiency_stats(
        dataset_name=dataset_name,
        num_batches=len(valid_dataloader),
        cfg_batch_size=config_each['Eval']['loader'].get('batch_size_per_card', None),
        model=model,
    )

    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        prev_loop_end = time.perf_counter()
        data_iter = iter(valid_dataloader)
        for _batch_idx in range(len(valid_dataloader)):
            t_fetch_start = prev_loop_end
            batch = next(data_iter)
            t_fetch_end = time.perf_counter()

            t_prep_start = time.perf_counter()
            batch_tensor, batch_numpy = split_batch_tensors(batch, device, max_tensor_items=5)
            raw_images = batch[3] if len(batch) > 3 else None
            t_prep_end = time.perf_counter()

            maybe_profile_first_batch(eff_stats, model, batch_tensor, device)
            sync_device(device)
            t_model_start = time.perf_counter()
            preds = model(batch_tensor[0], data=batch_tensor[1:])
            sync_device(device)
            t_model_end = time.perf_counter()

            post_result = post_process(preds, select_postprocess_batch(post_process, batch_numpy))
            selected_result = post_result
            if use_gtc_decode and isinstance(post_result, list) and len(post_result) == 2:
                selected_result = post_result[0] if infer_branch == 'gtc' else post_result[1]

            if isinstance(selected_result, tuple):
                texts, gts = selected_result
            else:
                texts, gts = selected_result, None

            sample_records = []
            for i, (txt, prob) in enumerate(texts):
                gt_text = ''
                if gts is not None and i < len(gts):
                    gt_text = gts[i][0]
                txt_norm = replace_punctuation(txt)
                gt_norm = replace_punctuation(gt_text)
                ned = 1 - Levenshtein.normalized_distance(txt_norm, gt_norm) if gt_norm is not None else 0.0
                cer, wer = update_htr_error_stats(htr_stats, txt_norm, gt_norm)
                ned_list.append(ned)
                num += 1
                if int(ned) == 1:
                    true_num += 1
                sample_records.append({
                    'img_name': f'{dataset_name}_{sample_offset + i}',
                    'type': dataset_name,
                    'label': gt_norm,
                    'pred': txt_norm,
                    'NED': float(ned),
                    'CER': float(cer),
                    'WER': float(wer),
                    'sample_img': raw_images[i] if raw_images is not None else None,
                })
            t_post_end = time.perf_counter()

            for record in sample_records:
                output_log['img_name'].append(record['img_name'])
                output_log['type'].append(record['type'])
                output_log['label'].append(record['label'])
                output_log['pred'].append(record['pred'])
                output_log['NED'].append(record['NED'])
                output_log['CER'].append(record['CER'])
                output_log['WER'].append(record['WER'])
                try:
                    image_bytes.append(to_png_bytes(record['sample_img']))
                except Exception:
                    image_bytes.append(None)
            t_dump_end = time.perf_counter()

            update_efficiency_stats(
                eff_stats,
                num_samples=len(sample_records),
                data_time=(t_fetch_end - t_fetch_start) + (t_prep_end - t_prep_start),
                model_time=t_model_end - t_model_start,
                post_time=t_post_end - t_model_end,
                dump_time=t_dump_end - t_post_end,
            )
            sample_offset += len(texts)
            pbar.update(1)
            prev_loop_end = t_dump_end
        pbar.close()

    model.train()
    pnacc = true_num / num if num else 0.0
    ned_mean = float(np.mean(ned_list)) if ned_list else 0.0
    return pnacc, ned_mean, num, eff_stats, htr_stats


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)
    FLAGS = vars(FLAGS)
    infer_branch = FLAGS.get('infer_branch', 'ctc')
    opt = FLAGS.pop('opt')
    cfg.merge_dict(FLAGS)
    cfg.merge_dict(opt)
    cfg = prepare_cfg(cfg, infer_branch=infer_branch)

    save_pred_xlsx = FLAGS.get('save_pred_xlsx')
    if save_pred_xlsx is None:
        save_pred_xlsx = os.path.join(cfg.cfg['Global']['output_dir'], 'preds_dump.xlsx')
    os.makedirs(os.path.dirname(save_pred_xlsx), exist_ok=True)

    trainer = Trainer(cfg, mode='eval')
    if infer_branch == 'gtc' and replace_gtc_decoder_with_gtc_only(trainer.model):
        trainer.logger.info('Use GTC-only decoder for AR/SMTR FPS evaluation.')
    trainer.logger.info(f'Inference branch: {infer_branch}')

    data_dirs_list = []
    if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
        data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
    else:
        data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
        if data_dir_single:
            data_dirs_list = [[data_dir_single]]

    # Optional custom override example (keep commented):
    data_dirs_list = [[
        r'/a800data1/lirunrui/lmdb_output/test/IAM_test/',
    ]]
    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label', []),
        ('pred', []),
        ('NED', []),
        ('CER', []),
        ('WER', []),
    ])
    image_bytes = []
    every_PNacc_list = []
    every_ned_list = []
    every_cer_list = []
    every_wer_list = []
    total_num = 0
    total_True_num = 0
    total_ned_list = []
    total_eff_stats = None
    total_htr_stats = None
    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            pnacc, ned_mean, num, eff_stats, htr_stats = dump_predictions_with_efficiency(
                trainer,
                datadir,
                output_log,
                dataset_name,
                image_bytes,
                infer_branch=infer_branch,
            )
            htr_metrics = finalize_htr_error_stats(htr_stats)
            print(
                f"{dataset_name}:\t\t acc: {100 * pnacc:6g}, "
                f"norm_edit_dis:{100 * ned_mean:6g}, "
                f"CER:{100 * htr_metrics['cer']:6g}, "
                f"WER:{100 * htr_metrics['wer']:6g}"
            )
            print_efficiency_summary(finalize_efficiency_stats(eff_stats))
            every_PNacc_list.append(pnacc)
            every_ned_list.append(ned_mean)
            every_cer_list.append(htr_metrics['cer'])
            every_wer_list.append(htr_metrics['wer'])
            total_num += num
            total_True_num += int(pnacc * num)
            total_ned_list.extend([ned_mean] * num)
            total_eff_stats = merge_efficiency_stats(total_eff_stats, eff_stats)
            total_htr_stats = merge_htr_error_stats(total_htr_stats, htr_stats)

    try:
        import pandas as pd
        df = pd.DataFrame(output_log)
        df.to_excel(save_pred_xlsx, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.utils import get_column_letter

            wb = load_workbook(save_pred_xlsx)
            ws = wb.active
            img_col = ws.max_column + 1
            ws.cell(row=1, column=img_col, value='image')
            img_col_letter = get_column_letter(img_col)

            target_px = 160
            ws.column_dimensions[img_col_letter].width = max(ws.column_dimensions[img_col_letter].width or 0, target_px / 7.0)

            embedded = 0
            for r_idx, data in enumerate(image_bytes, start=2):
                if not data:
                    continue
                img_obj = OpenpyxlImage(io.BytesIO(data))
                try:
                    with Image.open(io.BytesIO(data)) as pil_img:
                        w, h = pil_img.size
                except Exception:
                    w, h = None, None

                img_obj.width = target_px
                if w and h and w > 0:
                    img_obj.height = h * (target_px / float(w))
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, img_obj.height * 0.75)

                img_obj.anchor = f"{img_col_letter}{r_idx}"
                ws.add_image(img_obj)
                embedded += 1

            print(f"[INFO] Embedded {embedded} images into Excel")
            wb.save(save_pred_xlsx)
        except Exception as embed_err:
            print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")
        total_acc = (total_True_num / total_num) if total_num else 0.0
        total_ned = float(np.mean(total_ned_list)) if total_ned_list else 0.0
        total_htr_metrics = finalize_htr_error_stats(total_htr_stats or init_htr_error_stats())
        total_cer = total_htr_metrics['cer']
        total_wer = total_htr_metrics['wer']
        s_mean_acc = float(np.mean(every_PNacc_list)) if every_PNacc_list else 0.0
        s_mean_ned = float(np.mean(every_ned_list)) if every_ned_list else 0.0
        s_mean_cer = float(np.mean(every_cer_list)) if every_cer_list else 0.0
        s_mean_wer = float(np.mean(every_wer_list)) if every_wer_list else 0.0
        if every_PNacc_list:
            acc_arr = np.array(every_PNacc_list, dtype=np.float32)
            ned_arr = np.array(every_ned_list, dtype=np.float32)
            cer_arr = np.array(every_cer_list, dtype=np.float32)
            wer_arr = np.array(every_wer_list, dtype=np.float32)
            if len(acc_arr) == len(S_WEIGHT):
                weights = S_WEIGHT / np.sum(S_WEIGHT)
                s_weight_acc = float(np.sum(acc_arr * weights))
                s_weight_ned = float(np.sum(ned_arr * weights))
                s_weight_cer = float(np.sum(cer_arr * weights))
                s_weight_wer = float(np.sum(wer_arr * weights))
            else:
                print(f"[WARN] S_WEIGHT length mismatch: metrics={len(acc_arr)}, weights={len(S_WEIGHT)}. Fallback to S_mean.")
                s_weight_acc = s_mean_acc
                s_weight_ned = s_mean_ned
                s_weight_cer = s_mean_cer
                s_weight_wer = s_mean_wer
        else:
            s_weight_acc = 0.0
            s_weight_ned = 0.0
            s_weight_cer = 0.0
            s_weight_wer = 0.0
        print(
            f"total:\t\t acc: {100 * total_acc:6g}, "
            f"norm_edit_dis:{100 * total_ned:6g}, "
            f"CER:{100 * total_cer:6g}, WER:{100 * total_wer:6g}"
        )
        print(
            f"S_mean:\t\t acc: {100 * s_mean_acc:6g}, "
            f"norm_edit_dis:{100 * s_mean_ned:6g}, "
            f"CER:{100 * s_mean_cer:6g}, WER:{100 * s_mean_wer:6g}"
        )
        print(
            f"S_weight:\t\t acc: {100 * s_weight_acc:6g}, "
            f"norm_edit_dis:{100 * s_weight_ned:6g}, "
            f"CER:{100 * s_weight_cer:6g}, WER:{100 * s_weight_wer:6g}"
        )
        if total_eff_stats is not None:
            print_efficiency_summary(finalize_efficiency_stats(total_eff_stats))
        print(f'Predictions (with NED/CER/WER) saved to {save_pred_xlsx}')
    except Exception as e:
        print(f'[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.')


if __name__ == '__main__':
    main()
