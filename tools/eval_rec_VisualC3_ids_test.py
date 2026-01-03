import io
import os
import sys
import torch
import numpy as np
from collections import OrderedDict
from rapidfuzz.distance import Levenshtein
from tqdm import tqdm
from PIL import Image

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

from tools.data import build_dataloader
from tools.engine.config import Config
from tools.engine.trainer import Trainer
from tools.utility import ArgsParser


IDS_DICT_PATH = os.path.join(__dir__, 'utils', 'dict', 'visual_c3_ids', 'char_to_ids.txt')


# =========================
# 对齐：用 Levenshtein.opcodes 做全局对齐（避免 pad/truncate 吞 FP / 避免错位）
# =========================
def align_by_opcodes(gt: str, pred: str, gap_char=None):
    """
    将 gt/pred 对齐成“列对齐”序列，支持插入/删除/替换。
    返回列表，每个元素：(gt_ch or gap, pred_ch or gap, gt_idx or None)

    - gt_idx: 该列对应的“原始 gt 索引”，若该列是 pred 插入（gt 无字符）则为 None
    - gap_char: gap 占位符（默认 None）
    """
    aligned = []

    # opcodes: (tag, i1, i2, j1, j2), tag in {equal, replace, insert, delete} :contentReference[oaicite:2]{index=2}
    for tag, i1, i2, j1, j2 in Levenshtein.opcodes(gt, pred):
        if tag in ("equal", "replace"):
            len_a = i2 - i1
            len_b = j2 - j1
            m = min(len_a, len_b)

            # 对齐的主体部分（1-1）
            for k in range(m):
                gi = i1 + k
                pj = j1 + k
                aligned.append((gt[gi], pred[pj], gi))

            # replace 段里若长度不等，把剩余部分视为 delete/insert
            if len_a > m:
                for gi in range(i1 + m, i2):
                    aligned.append((gt[gi], gap_char, gi))
            if len_b > m:
                for pj in range(j1 + m, j2):
                    aligned.append((gap_char, pred[pj], None))

        elif tag == "delete":
            for gi in range(i1, i2):
                aligned.append((gt[gi], gap_char, gi))

        elif tag == "insert":
            for pj in range(j1, j2):
                aligned.append((gap_char, pred[pj], None))

        else:
            raise ValueError(f"Unknown opcode tag: {tag}")

    return aligned


# =========================
# 精简检测指标：混合（有错+无错）测试集最关键的指标集
# =========================
def _safe_div(num, den):
    return None if den == 0 else (num / den)

def _safe_pct(num, den):
    v = _safe_div(num, den)
    return None if v is None else (v * 100.0)

def calculate_cuo_metric_compact(gt_sentences, pred_sentences, X='X'):
    """
    仅评 X（错字位点）的检测，输出精简但高信息量的指标：
    - N_clean_sent: GT 无 X 的句子数
    - N_error_sent: GT 有 X 的句子数
    - Char_P / Char_R / Char_F1:
        precision = TP/(TP+FP), recall = TP/(TP+FN) :contentReference[oaicite:3]{index=3}
    - Sent_FA: clean sentence 上预测出任意 X 的比例（误报率）
    - Sent_EM: error sentence 上 exact match（X位置集合完全一致且不能多报/插入X）的比例
    """
    n = min(len(gt_sentences), len(pred_sentences))
    if n == 0:
        return {}

    n_clean_sent = 0
    n_error_sent = 0
    sent_fa = 0   # false alarm on clean sentences
    sent_em = 0   # exact match on error sentences

    # char-level TP/FP/FN（只针对 X）
    tp = 0
    fp = 0
    fn = 0

    for gt, pred in zip(gt_sentences[:n], pred_sentences[:n]):
        aligned = align_by_opcodes(gt, pred, gap_char=None)

        gt_x_pos = set()
        pred_x_pos = set()
        ins_x_cnt = 0

        for gch, pch, gi in aligned:
            g_is_x = (gch == X)
            p_is_x = (pch == X)

            # 统计位置集合（以 GT 坐标为准）
            if gi is not None and g_is_x:
                gt_x_pos.add(gi)

            if p_is_x:
                if gi is None:
                    # pred 在 GT gap（插入列）上输出 X -> 一定是多报(FP) + exact match 必失败
                    ins_x_cnt += 1
                else:
                    pred_x_pos.add(gi)

            # char-level 计数（仅 X 类）
            if gi is None:
                # 插入列：只要插入的是 X，就算 FP
                if p_is_x:
                    fp += 1
            else:
                if g_is_x and p_is_x:
                    tp += 1
                elif (not g_is_x) and p_is_x:
                    fp += 1
                elif g_is_x and (not p_is_x):
                    fn += 1

        gt_has_x = (len(gt_x_pos) > 0)
        pred_has_x = (len(pred_x_pos) > 0) or (ins_x_cnt > 0)

        if gt_has_x:
            n_error_sent += 1
            # exact match：X位置集合完全一致，且不能多报（含插入列X）
            if pred_has_x and (ins_x_cnt == 0) and (pred_x_pos == gt_x_pos):
                sent_em += 1
        else:
            n_clean_sent += 1
            # clean sentence 上只要预测出任意 X 就算误报
            if pred_has_x:
                sent_fa += 1

    # char-level metrics
    char_p = _safe_pct(tp, tp + fp)   # TP/(TP+FP) :contentReference[oaicite:4]{index=4}
    char_r = _safe_pct(tp, tp + fn)   # TP/(TP+FN) :contentReference[oaicite:5]{index=5}
    char_f1 = None
    if char_p is not None and char_r is not None and (char_p + char_r) > 0:
        char_f1 = 2 * char_p * char_r / (char_p + char_r)

    # sentence-level metrics
    sent_fa_rate = _safe_pct(sent_fa, n_clean_sent)  # clean 误报率
    sent_em_rate = _safe_pct(sent_em, n_error_sent)  # error exact-match 率

    return {
        "N_sent": n,
        "N_clean_sent": n_clean_sent,
        "N_error_sent": n_error_sent,
        "Char_P": char_p,
        "Char_R": char_r,
        "Char_F1": char_f1,
        "Sent_FA": sent_fa_rate,
        "Sent_EM": sent_em_rate,
        # 下面三项只为方便定位（可不打印）
        "Char_TP": tp,
        "Char_FP": fp,
        "Char_FN": fn,
    }


# =========================
# 映射与文本处理
# =========================
def load_ids_mapping(path: str):
    """加载 char->ids 映射，并构建 ids->char 反查表。"""
    char2ids = {}
    ids2char = {}
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(maxsplit=1)
            if len(parts) != 2:
                continue
            ch, ids_seq = parts
            char2ids[ch] = ids_seq
            # 若多个字符映射到同一 IDS，保留最先出现的
            ids2char.setdefault(ids_seq, ch)
    return char2ids, ids2char


def ids_seq_to_text(ids_seq: str, ids2char: dict, unknown_char: str = 'X') -> str:
    """把以空格分隔的 IDS 序列还原为字符串；缺失用 unknown_char。"""
    if ids_seq is None:
        return ''
    tokens = ids_seq.strip().split()
    chars = [ids2char.get(tok, unknown_char) for tok in tokens]
    return ''.join(chars)


def replace_punctuation(text: str) -> str:
    """与原脚本保持一致的标点替换。"""
    if text is None:
        return ''
    mapping = {
        r'，': r',', r'。': r'.', r'！': r'!', r'？': r'?', r'；': r';', r'：': r':',
        r'“': r'"', r'”': r'"', r'‘': r"'", r'’': r"'",
    }
    for k, v in mapping.items():
        text = text.replace(k, v)
    return text


def parse_args():
    parser = ArgsParser()
    parser.add_argument(
        '--save_pred_xlsx',
        type=str,
        default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump.xlsx',
    )
    parser.add_argument(
        '--ids_dict_path',
        type=str,
        default=IDS_DICT_PATH,
        help='Path to char_to_ids mapping file.',
    )
    args = parser.parse_args()
    return args


def prepare_cfg(cfg):
    if cfg.cfg['Global']['output_dir'][-1] == '/':
        cfg.cfg['Global']['output_dir'] = cfg.cfg['Global']['output_dir'][:-1]
    if cfg.cfg['Global']['pretrained_model'] is None:
        cfg.cfg['Global']['pretrained_model'] = cfg.cfg['Global']['output_dir'] + '/best.pth'
    cfg.cfg['Global']['use_amp'] = False
    cfg.cfg['PostProcess']['with_ratio'] = True
    cfg.cfg['Metric']['with_ratio'] = True
    cfg.cfg['Metric']['max_len'] = 100
    cfg.cfg['Global']['max_text_length'] = 100
    cfg.cfg['Metric']['max_ratio'] = 12
    keep_keys = cfg.cfg['Eval']['dataset']['transforms'][-1]['KeepKeys']['keep_keys']
    if 'real_ratio' not in keep_keys:
        keep_keys.append('real_ratio')
    return cfg

def split_src_tgt(gt_raw: str):
    """
    解析 faked 标签：'src | | | tgt' -> (src, tgt)
    兼容：
      - 没有 '| | |'：只返回 (gt_raw, None)
      - 有 '| | |' ：返回 (src, tgt)
    """
    if gt_raw is None:
        return "", None
    s = str(gt_raw).strip()
    if '? ? ?' not in s:
        return s, None
    parts = s.split('? ? ?')
    
    if len(parts) != 2:
        print(f"[WARN] Unexpected faked gt format: {gt_raw}")
        return s, None
    
    # print(f"[INFO] Splitting faked gt: {parts}") # debug
    src = parts[0].strip()
    tgt = parts[1].strip()
    
    return src, tgt


def to_png_bytes(img_array):
    """将 numpy 图像数组转成 PNG bytes，兼容单通道/三通道及 0~1/0~255 输入。"""
    if img_array is None:
        return None

    arr = np.array(img_array)
    if arr.ndim == 3 and arr.shape[0] in (1, 3):
        arr = np.transpose(arr, (1, 2, 0))
    arr = np.squeeze(arr)

    # 归一化到 0-255
    if arr.max() <= 1.0 + 1e-3:
        arr = arr * 255.0
    arr = np.clip(arr, 0, 255).astype(np.uint8)

    mode = 'L' if arr.ndim == 2 else 'RGB'
    img = Image.fromarray(arr, mode=mode)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    buf.close()
    return data

def dump_predictions(trainer, datadir, output_log, dataset_name, ids2char, det_inputs, image_bytes):
    config_each = trainer.cfg.copy()
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    model.eval()
    num = 0
    true_num = 0
    ned_list = []
    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        for batch_idx, batch in enumerate(valid_dataloader):
            # batch: [image_tensor, label_tensor, length_tensor, raw_images(list/np)]
            batch_tensor = [t.to(device) for t in batch[:3]]
            batch_numpy = [t.numpy() for t in batch[:3]]
            raw_images = batch[3] if len(batch) > 3 else None
            preds = model(batch_tensor[0], data=batch_tensor[1:])
            post_result = post_process(preds, batch_numpy)
            if isinstance(post_result, tuple):
                texts, gts = post_result
            else:
                texts, gts = post_result, None

            for i, (txt_ids, prob) in enumerate(texts):
                gt_ids = ''
                if gts is not None and i < len(gts):
                    gt_ids = gts[i][0]

                # ===== 关键修改：解析 'src|||tgt'，先只用 src检测, 纠错这块还没想好。 =====
                gt_src_ids, gt_tgt_ids = split_src_tgt(gt_ids)
                gt_ids = gt_src_ids    

                # IDS -> 字符
                txt = ids_seq_to_text(txt_ids, ids2char, unknown_char='X')
                gt_text = ids_seq_to_text(gt_ids, ids2char, unknown_char='?')  # 应该是不会存在未知 IDS 的

                txt_norm = replace_punctuation(txt.strip())
                gt_norm = replace_punctuation(gt_text.strip())

                # 收集“X 检测”所需序列（gt_norm 应是含 X 的 src label；pred 是模型输出含/不含X）
                det_inputs['gts'].append(gt_norm)
                det_inputs['preds'].append(txt_norm)

                # NED（字符串相似度）
                ned = 1 - Levenshtein.normalized_distance(txt_norm, gt_norm) if gt_norm is not None else 0.0
                ned_list.append(ned)
                num += 1
                if int(ned) == 1:
                    true_num += 1

                img_name = f"{dataset_name}_{sample_offset + i}"
                output_log['img_name'].append(img_name)
                output_log['type'].append(dataset_name)
                output_log['label'].append(gt_norm)
                output_log['pred'].append(txt_norm)
                output_log['label_ids'].append(str(gt_ids))
                output_log['pred_ids'].append(str(txt_ids))
                output_log['NED'].append(float(ned))

                # 还原当前样本的图像（经过 transform），以便后续写入 XLSX
                try:
                    sample_img = raw_images[i] if raw_images is not None else None
                    image_bytes.append(to_png_bytes(sample_img))
                except Exception:
                    image_bytes.append(None)

            sample_offset += len(texts)
            pbar.update(1)
        pbar.close()

    model.train()
    pnacc = true_num / num if num else 0.0
    ned_mean = float(np.mean(ned_list)) if ned_list else 0.0
    return pnacc, ned_mean, num


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)
    FLAGS = vars(FLAGS)
    opt = FLAGS.pop('opt')
    cfg.merge_dict(FLAGS)
    cfg.merge_dict(opt)
    cfg = prepare_cfg(cfg)

    # 加载 IDS 映射
    ids_dict_path = FLAGS.get('ids_dict_path') or IDS_DICT_PATH
    _, ids2char = load_ids_mapping(ids_dict_path)

    save_pred_xlsx = FLAGS.get('save_pred_xlsx')
    if save_pred_xlsx is None:
        save_pred_xlsx = os.path.join(cfg.cfg['Global']['output_dir'], 'preds_dump_ids.xlsx')
    os.makedirs(os.path.dirname(save_pred_xlsx), exist_ok=True)

    trainer = Trainer(cfg, mode='eval')

    data_dirs_list = []
    if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
        data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
    else:
        data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
        if data_dir_single:
            data_dirs_list = [[data_dir_single]]

    # 可按需覆盖
    data_dirs_list = [[
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_ids/test_ids_lmdb/test_correct',
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_ids/test_ids_lmdb/test_faked'
    ]]

    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label', []),
        ('pred', []),
        ('label_ids', []),
        ('pred_ids', []),
        ('NED', []),
    ])
    det_inputs = {'gts': [], 'preds': []}
    image_bytes = []

    every_PNacc_list = []
    every_ned_list = []
    total_num = 0
    total_True_num = 0
    total_ned_list = []

    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            pnacc, ned_mean, num = dump_predictions(trainer, datadir, output_log, dataset_name, ids2char, det_inputs, image_bytes)
            print(f"{dataset_name}:\t\t acc: {100 * pnacc:6g}, norm_edit_dis:{100 * ned_mean:6g}")
            every_PNacc_list.append(pnacc)
            every_ned_list.append(ned_mean)
            total_num += num
            total_True_num += int(pnacc * num)
            total_ned_list.extend([ned_mean] * num)

    try:
        import pandas as pd
        df = pd.DataFrame(output_log)
        df.to_excel(save_pred_xlsx, index=False)

        # 将图像嵌入到 Excel（列 F，标题 image）。需要 openpyxl。
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.utils import get_column_letter

            wb = load_workbook(save_pred_xlsx)
            ws = wb.active
            img_col = ws.max_column + 1
            ws.cell(row=1, column=img_col, value='image')
            img_col_letter = get_column_letter(img_col)

            # 设定列宽以容纳缩略图（约 160px）
            target_px = 160
            ws.column_dimensions[img_col_letter].width = max(ws.column_dimensions[img_col_letter].width or 0, target_px / 7.0)

            embedded = 0
            for r_idx, data in enumerate(image_bytes, start=2):
                if not data:
                    continue
                img_obj = OpenpyxlImage(io.BytesIO(data))

                # 先用 PIL 读取尺寸，避免访问私有属性
                try:
                    from PIL import Image as PILImage
                    with PILImage.open(io.BytesIO(data)) as pil_img:
                        w, h = pil_img.size
                except Exception:
                    w, h = None, None

                img_obj.width = target_px
                if w and h and w > 0:
                    img_obj.height = h * (target_px / float(w))
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, img_obj.height * 0.75)

                # 直接锚定到目标单元格坐标，依赖行高/列宽约束尺寸
                img_obj.anchor = f"{img_col_letter}{r_idx}"
                ws.add_image(img_obj)
                embedded += 1

            print(f"[INFO] Embedded {embedded} images into Excel")

            wb.save(save_pred_xlsx)
        except Exception as embed_err:
            print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")
        total_acc = (total_True_num / total_num) if total_num else 0.0
        total_ned = float(np.mean(total_ned_list)) if total_ned_list else 0.0
        s_mean_acc = float(np.mean(every_PNacc_list)) if every_PNacc_list else 0.0
        s_mean_ned = float(np.mean(every_ned_list)) if every_ned_list else 0.0
        s_weight_acc = float(np.sum(np.array(every_PNacc_list))) if every_PNacc_list else 0.0
        s_weight_ned = float(np.sum(np.array(every_ned_list))) if every_ned_list else 0.0
        print(f"total:\t\t acc: {100 * total_acc:6g}, norm_edit_dis:{100 * total_ned:6g}")
        print(f"S_mean:\t\t acc: {100 * s_mean_acc:6g}, norm_edit_dis:{100 * s_mean_ned:6g}")
        print(f"S_weight:\t\t acc: {100 * s_weight_acc:6g}, norm_edit_dis:{100 * s_weight_ned:6g}")
        print(f'Predictions (with NED) saved to {save_pred_xlsx}')
    except Exception as e:
        print(f'[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.')

    # ========= 精简的 X 检测指标 =========
    det = calculate_cuo_metric_compact(det_inputs['gts'], det_inputs['preds'], X='X')

    def fmt(x):
        return "N/A" if x is None else f"{x:.3f}"

    print("\nCuo detection metrics (compact, mixed clean+error):")
    print(f"N_sent={det['N_sent']} | clean={det['N_clean_sent']} | error={det['N_error_sent']}")
    print(f"Char_P={fmt(det['Char_P'])}%  Char_R={fmt(det['Char_R'])}%  Char_F1={fmt(det['Char_F1'])}%")
    print(f"Sent_FA={fmt(det['Sent_FA'])}%  Sent_EM={fmt(det['Sent_EM'])}%")
    # 如需排查，可临时打开这一行（不建议默认打印太多）
    # print(f"(debug) Char_TP={det['Char_TP']} Char_FP={det['Char_FP']} Char_FN={det['Char_FN']}")


if __name__ == '__main__':
    main()
