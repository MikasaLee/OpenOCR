import copy
import io
import json
import os
import sys
import time
from collections import OrderedDict

import numpy as np
import torch
from PIL import Image
from tqdm import tqdm

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

import eval_rec_textline_triple as triple_base

from tools.data import build_dataloader
from tools.engine.config import Config
from tools.engine.trainer import Trainer
from tools.utility import ArgsParser
from tools.utils.ids_syntax import DEFAULT_IDC_ARITY, validate_ids_prefix


def str2bool(v):
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ('1', 'true', 'yes', 'y', 'on'):
        return True
    if s in ('0', 'false', 'no', 'n', 'off'):
        return False
    raise ValueError(f'Cannot parse boolean value from: {v}')


def split_batch(batch, device, max_tensor_items=6):
    tensor_items = []
    extra_items = []
    tensor_block = True
    for item in batch:
        if tensor_block and torch.is_tensor(item) and len(tensor_items) < max_tensor_items:
            tensor_items.append(item)
        else:
            tensor_block = False
            extra_items.append(item)
    batch_tensor = [t.to(device) for t in tensor_items]
    batch_numpy = [t.numpy() for t in tensor_items]
    return batch_tensor, batch_numpy, extra_items


def find_raw_images(extra_fields):
    for item in extra_fields:
        if isinstance(item, list) and len(item) > 0:
            first = item[0]
            if isinstance(first, np.ndarray) and first.ndim >= 2:
                return item
            if isinstance(first, bytes):
                return item
    return None


def sync_device(device):
    if not torch.cuda.is_available():
        return
    try:
        torch.cuda.synchronize(device)
    except Exception:
        torch.cuda.synchronize()


def norm_pairs(lst):
    out = []
    for item in lst:
        if isinstance(item, (tuple, list)) and len(item) >= 1:
            s = item[0]
            p = item[1] if len(item) > 1 else 1.0
            out.append((str(s), float(p)))
        else:
            out.append((str(item), 1.0))
    return out


def build_eval_dataloader_from_dir(trainer, datadir, batch_size_override=None):
    config_each = copy.deepcopy(trainer.cfg)
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir

    if batch_size_override is not None:
        config_each['Eval']['loader']['batch_size_per_card'] = int(batch_size_override)
        config_each['Eval']['loader']['drop_last'] = False
        config_each['Eval']['loader']['shuffle'] = False

    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')
    return config_each, valid_dataloader


def parse_post_result(post_result, post_process):
    gts = None
    outputs = None
    if isinstance(post_result, dict):
        outputs = post_result
        gts = post_result.get('label_text', None)
    elif isinstance(post_result, list) and len(post_result) == 2 and isinstance(post_result[0], tuple):
        (text_res, text_gt), (ids_res, _ids_gt) = post_result
        ids_text_res = []
        for item in ids_res:
            ids_str = item[0] if isinstance(item, (tuple, list)) else item
            conf = item[1] if isinstance(item, (tuple, list)) and len(item) > 1 else 1.0
            if hasattr(post_process, 'map_ids_to_text'):
                ids_text_res.append((post_process.map_ids_to_text(ids_str), conf))
            else:
                ids_text_res.append(('', conf))
        outputs = {
            'text': text_res,
            'ids': ids_res,
            'text_from_ids': ids_text_res,
        }
        gts = text_gt
    elif isinstance(post_result, tuple):
        outputs, gts = post_result
        if not isinstance(outputs, dict):
            outputs = {'text': outputs}
    else:
        outputs = {'text': post_result}
    return outputs, gts


def init_efficiency_stats(dataset_name, num_batches, cfg_batch_size):
    return {
        'dataset_name': dataset_name,
        'num_batches': int(num_batches),
        'cfg_batch_size': int(cfg_batch_size) if cfg_batch_size is not None else None,
        'num_samples': 0,
        'total_data': 0.0,
        'total_model': 0.0,
        'total_post': 0.0,
        'total_dump': 0.0,
    }


def finalize_efficiency_stats(stats):
    summary = dict(stats)
    num_batches = summary['num_batches']
    num_samples = summary['num_samples']

    summary['total_eval'] = summary['total_data'] + summary['total_model'] + summary['total_post']
    summary['total_loop'] = summary['total_eval'] + summary['total_dump']
    summary['avg_batch_size'] = triple_base._safe_div(num_samples, num_batches)
    summary['throughput'] = triple_base._safe_div(num_samples, summary['total_eval'])
    summary['throughput_with_dump'] = triple_base._safe_div(num_samples, summary['total_loop'])
    summary['avg_batch_e2e'] = triple_base._safe_div(summary['total_eval'], num_batches)
    summary['avg_batch_loop'] = triple_base._safe_div(summary['total_loop'], num_batches)
    summary['avg_batch_data'] = triple_base._safe_div(summary['total_data'], num_batches)
    summary['avg_batch_model'] = triple_base._safe_div(summary['total_model'], num_batches)
    summary['avg_batch_post'] = triple_base._safe_div(summary['total_post'], num_batches)
    summary['avg_batch_dump'] = triple_base._safe_div(summary['total_dump'], num_batches)
    summary['avg_model_sample_ms'] = None if num_samples == 0 else (summary['total_model'] / num_samples * 1000.0)
    summary['avg_post_sample_ms'] = None if num_samples == 0 else (summary['total_post'] / num_samples * 1000.0)
    return summary


def merge_efficiency_stats(dst, src):
    if dst is None:
        merged = init_efficiency_stats('total', src['num_batches'], src.get('cfg_batch_size'))
        merged['num_samples'] = src['num_samples']
        merged['total_data'] = src['total_data']
        merged['total_model'] = src['total_model']
        merged['total_post'] = src['total_post']
        merged['total_dump'] = src['total_dump']
        return merged

    dst['num_batches'] += src['num_batches']
    dst['num_samples'] += src['num_samples']
    dst['total_data'] += src['total_data']
    dst['total_model'] += src['total_model']
    dst['total_post'] += src['total_post']
    dst['total_dump'] += src['total_dump']
    if dst.get('cfg_batch_size') != src.get('cfg_batch_size'):
        dst['cfg_batch_size'] = None
    return dst


def print_efficiency_summary(summary, latency=None):
    dataset_name = summary['dataset_name']
    avg_bs = 'N/A' if summary['avg_batch_size'] is None else f"{summary['avg_batch_size']:.2f}"
    cfg_bs = 'N/A' if summary['cfg_batch_size'] is None else str(summary['cfg_batch_size'])
    throughput = 'N/A' if summary['throughput'] is None else f"{summary['throughput']:.2f}"
    throughput_with_dump = 'N/A' if summary['throughput_with_dump'] is None else f"{summary['throughput_with_dump']:.2f}"
    avg_model_ms = 'N/A' if summary['avg_model_sample_ms'] is None else f"{summary['avg_model_sample_ms']:.2f}"
    avg_post_ms = 'N/A' if summary['avg_post_sample_ms'] is None else f"{summary['avg_post_sample_ms']:.2f}"

    print(f"[EFF][{dataset_name}] samples={summary['num_samples']}, batches={summary['num_batches']}, cfg_bs={cfg_bs}, avg_bs={avg_bs}")
    print(f"[EFF][{dataset_name}] end2end_total={summary['total_eval']:.4f}s, throughput={throughput} lines/s, avg_batch_e2e={summary['avg_batch_e2e']:.4f}s")
    print(f"[EFF][{dataset_name}] loop_total={summary['total_loop']:.4f}s, throughput_with_dump={throughput_with_dump} lines/s, avg_batch_loop={summary['avg_batch_loop']:.4f}s")
    print(
        f"[EFF][{dataset_name}] avg_batch_data={summary['avg_batch_data']:.4f}s, "
        f"avg_batch_model={summary['avg_batch_model']:.4f}s, "
        f"avg_batch_post={summary['avg_batch_post']:.4f}s, "
        f"avg_batch_dump={summary['avg_batch_dump']:.4f}s"
    )
    print(f"[EFF][{dataset_name}] model_sample_time={avg_model_ms} ms/line, post_sample_time={avg_post_ms} ms/line")

    if latency is not None:
        latency_avg = 'N/A' if latency.get('avg_ms') is None else f"{latency['avg_ms']:.2f}"
        latency_p50 = 'N/A' if latency.get('p50_ms') is None else f"{latency['p50_ms']:.2f}"
        latency_p90 = 'N/A' if latency.get('p90_ms') is None else f"{latency['p90_ms']:.2f}"
        fixed_first_n = latency.get('fixed_first_n_samples', None)
        fixed_first_n_str = '' if fixed_first_n is None else f", fixed_first_n={fixed_first_n}"
        print(
            f"[EFF][{dataset_name}] latency_bs1_model={latency_avg} ms/line, "
            f"p50={latency_p50} ms, p90={latency_p90} ms, "
            f"warmup={latency['warmup']}, measured={latency['num_measured']}{fixed_first_n_str}"
        )


def aggregate_latency_results(latency_results):
    valid = []
    for item in latency_results.values():
        latency = item.get('latency_bs1', item) if isinstance(item, dict) else item
        if latency is None:
            continue
        if latency.get('avg_ms') is None or latency.get('num_measured', 0) <= 0:
            continue
        valid.append(latency)
    if not valid:
        return None

    total_measured = sum(item['num_measured'] for item in valid)
    weighted_avg = sum(item['avg_ms'] * item['num_measured'] for item in valid) / total_measured
    return {
        'dataset_name': 'total',
        'avg_ms': float(weighted_avg),
        'p50_ms': None,
        'p90_ms': None,
        'num_measured': int(total_measured),
        'warmup': valid[0].get('warmup'),
        'fixed_first_n_samples': sum(item.get('fixed_first_n_samples', 0) for item in valid),
    }


def print_efficiency_table_summary(tag, summary, latency=None):
    cfg_bs = summary.get('cfg_batch_size')
    throughput_key = f"throughput_bs{cfg_bs}" if cfg_bs is not None else "throughput"
    latency_avg = 'N/A' if latency is None or latency.get('avg_ms') is None else f"{latency['avg_ms']:.2f}"
    throughput = 'N/A' if summary.get('throughput') is None else f"{summary['throughput']:.2f}"
    model_sample_time = 'N/A' if summary.get('avg_model_sample_ms') is None else f"{summary['avg_model_sample_ms']:.2f}"
    print(
        f"[EFF-SUMMARY][{tag}] "
        f"latency_bs1_model={latency_avg} ms/line | "
        f"{throughput_key}={throughput} lines/s | "
        f"model_sample_time={model_sample_time} ms/line"
    )


def parse_args():
    parser = ArgsParser()
    parser.add_argument(
        '--save_pred_xlsx',
        type=str,
        default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump_text_triple_eff.xlsx',
    )
    parser.add_argument(
        '--skip_xlsx',
        type=str2bool,
        default=False,
        help='Skip XLSX/image dump so efficiency runs are not polluted by dump artifacts.',
    )
    parser.add_argument(
        '--measure_latency_bs1',
        type=str2bool,
        default=False,
        help='Additionally measure bs=1 model-only latency with warmup.',
    )
    parser.add_argument(
        '--latency_warmup',
        type=int,
        default=20,
        help='Warmup iterations to exclude for bs=1 model-only latency.',
    )
    parser.add_argument(
        '--latency_max_samples',
        type=int,
        default=100,
        help='Maximum measured samples for bs=1 latency. <=0 means use all remaining samples after warmup.',
    )
    parser.add_argument(
        '--save_eff_json',
        type=str,
        default=None,
        help='Optional path to save efficiency summaries as JSON.',
    )
    parser.add_argument(
        '--use_cfg_eval_dirs',
        type=str2bool,
        default=False,
        help='Use Eval.dataset.data_dir/data_dir_list from config. Default False to match eval_rec_textline_triple.py fixed two-set behavior.',
    )
    return parser.parse_args()


def dump_predictions(
    trainer,
    datadir,
    output_log,
    dataset_name,
    image_bytes,
    det_inputs_text,
    det_inputs_idsText,
    idc_arity=None,
    collect_artifacts=True,
):
    config_each, valid_dataloader = build_eval_dataloader_from_dir(trainer, datadir)

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    model.eval()

    num = 0
    src_true_text = 0
    tgt_true_text = 0
    src_ned_text = []
    tgt_ned_text = []

    src_true_idsText = 0
    tgt_true_idsText = 0
    src_ned_idsText = []
    tgt_ned_idsText = []

    if idc_arity is None:
        idc_arity = DEFAULT_IDC_ARITY
    legal_token = 0
    total_token = 0
    legal_seq = 0
    total_seq = 0

    eff_stats = init_efficiency_stats(
        dataset_name=dataset_name,
        num_batches=len(valid_dataloader),
        cfg_batch_size=config_each['Eval']['loader'].get('batch_size_per_card', None),
    )

    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        prev_loop_end = time.perf_counter()
        data_iter = iter(valid_dataloader)

        for _batch_idx in range(len(valid_dataloader)):
            t_fetch_start = prev_loop_end
            batch = next(data_iter)
            t_fetch_end = time.perf_counter()

            t_prep_start = time.perf_counter()
            batch_tensor, batch_numpy, batch_extra = split_batch(batch, device, max_tensor_items=6)
            raw_images = find_raw_images(batch_extra)
            model_data = batch_tensor[1:]
            t_model_start = time.perf_counter()

            sync_device(device)
            preds = model(batch_tensor[0], data=model_data)
            sync_device(device)
            t_model_end = time.perf_counter()

            post_result = post_process(preds, batch_numpy)
            outputs, gts = parse_post_result(post_result, post_process)

            text_list = norm_pairs(outputs.get('text', []))
            ids_list = norm_pairs(outputs.get('ids', []))
            ids_text_list = norm_pairs(outputs.get('text_from_ids', []))

            sample_records = []
            for i in range(len(text_list)):
                pred_text, _ = text_list[i]
                pred_ids, _ = ids_list[i] if i < len(ids_list) else ('', 1.0)
                pred_ids_text, _ = ids_text_list[i] if i < len(ids_text_list) else ('', 1.0)

                gt_text = ''
                if gts is not None and i < len(gts):
                    gt_item = gts[i]
                    gt_text = gt_item[0] if isinstance(gt_item, (tuple, list)) and len(gt_item) >= 1 else gt_item
                if (gt_text is None or gt_text == '') and len(batch_numpy) > 1 and i < len(batch_numpy[1]):
                    try:
                        gt_text = batch_numpy[1][i].decode('utf-8') if isinstance(batch_numpy[1][i], bytes) else str(batch_numpy[1][i])
                    except Exception:
                        gt_text = str(batch_numpy[1][i])

                gt_src_txt, gt_tgt_txt = triple_base.split_src_tgt(gt_text)
                pred_text_norm = triple_base.replace_punctuation(pred_text)
                pred_ids_text_norm = triple_base.replace_punctuation(pred_ids_text)
                gt_src_norm = triple_base.replace_punctuation(gt_src_txt)
                gt_tgt_norm = triple_base.replace_punctuation(gt_tgt_txt)

                s_ned = 1 - triple_base.Levenshtein.normalized_distance(pred_text_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                t_ned = 1 - triple_base.Levenshtein.normalized_distance(pred_text_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_text.append(s_ned)
                tgt_ned_text.append(t_ned)
                if int(s_ned) == 1:
                    src_true_text += 1
                if int(t_ned) == 1:
                    tgt_true_text += 1

                s_ned_ids_text = 1 - triple_base.Levenshtein.normalized_distance(pred_ids_text_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                t_ned_ids_text = 1 - triple_base.Levenshtein.normalized_distance(pred_ids_text_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_idsText.append(s_ned_ids_text)
                tgt_ned_idsText.append(t_ned_ids_text)
                if int(s_ned_ids_text) == 1:
                    src_true_idsText += 1
                if int(t_ned_ids_text) == 1:
                    tgt_true_idsText += 1

                num += 1

                tokens = [tok for tok in str(pred_ids).strip().split() if tok]
                token_validity = []
                for tok in tokens:
                    ok, _, _ = validate_ids_prefix(
                        [ch for ch in tok if not ch.isspace()],
                        idc_arity=idc_arity,
                        require_closed=True,
                    )
                    token_validity.append(ok)
                is_sample_valid = all(token_validity)

                if triple_base.maybe_ids_tokens(tokens):
                    for ok in token_validity:
                        total_token += 1
                        if ok:
                            legal_token += 1
                    if tokens:
                        total_seq += 1
                        if is_sample_valid:
                            legal_seq += 1

                det_inputs_text['gts'].append(gt_src_norm)
                det_inputs_text['preds'].append(pred_text_norm)
                det_inputs_idsText['gts'].append(gt_src_norm)
                det_inputs_idsText['preds'].append(pred_ids_text_norm)

                sample_records.append({
                    'img_name': f"{dataset_name}_{sample_offset + i}",
                    'type': dataset_name,
                    'label_src': gt_src_norm,
                    'label_tgt': gt_tgt_norm,
                    'pred_text': pred_text_norm,
                    'pred_idsText': pred_ids_text_norm,
                    'pred_ids': str(pred_ids),
                    'ids_valid': is_sample_valid,
                    'NED_text_src': float(s_ned),
                    'NED_text_tgt': float(t_ned),
                    'NED_idsText_src': float(s_ned_ids_text),
                    'NED_idsText_tgt': float(t_ned_ids_text),
                    'sample_img': raw_images[i] if raw_images is not None and i < len(raw_images) else None,
                })

            t_post_end = time.perf_counter()

            if collect_artifacts:
                for record in sample_records:
                    output_log['img_name'].append(record['img_name'])
                    output_log['type'].append(record['type'])
                    output_log['label_src'].append(record['label_src'])
                    output_log['label_tgt'].append(record['label_tgt'])
                    output_log['pred_text'].append(record['pred_text'])
                    output_log['pred_idsText'].append(record['pred_idsText'])
                    output_log['pred_ids'].append(record['pred_ids'])
                    output_log['ids_valid'].append(record['ids_valid'])
                    output_log['NED_text_src'].append(record['NED_text_src'])
                    output_log['NED_text_tgt'].append(record['NED_text_tgt'])
                    output_log['NED_idsText_src'].append(record['NED_idsText_src'])
                    output_log['NED_idsText_tgt'].append(record['NED_idsText_tgt'])
                    try:
                        image_bytes.append(triple_base.to_png_bytes(record['sample_img']))
                    except Exception:
                        image_bytes.append(None)

            t_dump_end = time.perf_counter()

            batch_data_time = (t_fetch_end - t_fetch_start) + (t_model_start - t_prep_start)
            batch_model_time = t_model_end - t_model_start
            batch_post_time = t_post_end - t_model_end
            batch_dump_time = t_dump_end - t_post_end

            eff_stats['num_samples'] += len(sample_records)
            eff_stats['total_data'] += batch_data_time
            eff_stats['total_model'] += batch_model_time
            eff_stats['total_post'] += batch_post_time
            eff_stats['total_dump'] += batch_dump_time

            sample_offset += len(text_list)
            pbar.update(1)
            prev_loop_end = t_dump_end
        pbar.close()

    model.train()

    legality = {
        'legal_token': legal_token,
        'total_token': total_token,
        'legal_seq': legal_seq,
        'total_seq': total_seq,
    }

    return (
        src_true_text / num if num else 0.0,
        tgt_true_text / num if num else 0.0,
        float(np.mean(src_ned_text)) if src_ned_text else 0.0,
        float(np.mean(tgt_ned_text)) if tgt_ned_text else 0.0,
        src_true_idsText / num if num else 0.0,
        tgt_true_idsText / num if num else 0.0,
        float(np.mean(src_ned_idsText)) if src_ned_idsText else 0.0,
        float(np.mean(tgt_ned_idsText)) if tgt_ned_idsText else 0.0,
        num,
        legality,
        finalize_efficiency_stats(eff_stats),
    )


def measure_latency_bs1(trainer, datadir, dataset_name, warmup=20, max_samples=100):
    _config_each, valid_dataloader = build_eval_dataloader_from_dir(
        trainer,
        datadir,
        batch_size_override=1,
    )

    model = trainer.model
    device = trainer.device
    model.eval()

    latencies_ms = []
    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'latency {dataset_name}', position=0, leave=True)
        for batch_idx, batch in enumerate(valid_dataloader):
            batch_tensor, _batch_numpy, _batch_extra = split_batch(batch, device, max_tensor_items=6)
            model_data = batch_tensor[1:]

            sync_device(device)
            t0 = time.perf_counter()
            _ = model(batch_tensor[0], data=model_data)
            sync_device(device)
            t1 = time.perf_counter()

            if batch_idx >= warmup:
                latencies_ms.append((t1 - t0) * 1000.0)
                if max_samples > 0 and len(latencies_ms) >= max_samples:
                    pbar.update(1)
                    break
            pbar.update(1)
        pbar.close()

    model.train()

    if not latencies_ms:
        return {
            'dataset_name': dataset_name,
            'avg_ms': None,
            'p50_ms': None,
            'p90_ms': None,
            'num_measured': 0,
            'warmup': warmup,
            'fixed_first_n_samples': max(warmup, 0),
        }

    lat_arr = np.array(latencies_ms, dtype=np.float64)
    if max_samples > 0:
        fixed_first_n_samples = warmup + min(max_samples, int(lat_arr.size))
    else:
        fixed_first_n_samples = warmup + int(lat_arr.size)
    return {
        'dataset_name': dataset_name,
        'avg_ms': float(lat_arr.mean()),
        'p50_ms': float(np.percentile(lat_arr, 50)),
        'p90_ms': float(np.percentile(lat_arr, 90)),
        'num_measured': int(lat_arr.size),
        'warmup': warmup,
        'fixed_first_n_samples': int(fixed_first_n_samples),
    }


def maybe_save_xlsx(save_pred_xlsx, output_log, image_bytes):
    import pandas as pd

    df = pd.DataFrame(output_log)
    df.to_excel(save_pred_xlsx, index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.utils import get_column_letter

        wb = load_workbook(save_pred_xlsx)
        ws = wb.active
        img_col = ws.max_column + 1
        ws.cell(row=1, column=img_col, value='image')
        img_col_letter = get_column_letter(img_col)
        target_px = 160
        ws.column_dimensions[img_col_letter].width = max(
            ws.column_dimensions[img_col_letter].width or 0,
            target_px / 7.0,
        )
        embedded = 0
        for r_idx, data in enumerate(image_bytes, start=2):
            if not data:
                continue
            img_obj = OpenpyxlImage(io.BytesIO(data))
            try:
                with Image.open(io.BytesIO(data)) as pil_img:
                    w, h = pil_img.size
            except Exception:
                w, h = None, None
            img_obj.width = target_px
            if w and h and w > 0:
                img_obj.height = h * (target_px / float(w))
            ws.row_dimensions[r_idx].height = max(
                ws.row_dimensions[r_idx].height or 0,
                img_obj.height * 0.75,
            )
            img_obj.anchor = f"{img_col_letter}{r_idx}"
            ws.add_image(img_obj)
            embedded += 1
        print(f"[INFO] Embedded {embedded} images into Excel")
        wb.save(save_pred_xlsx)
    except Exception as embed_err:
        print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")

    print(f"Predictions saved to {save_pred_xlsx}")


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)

    flags = vars(FLAGS)
    opt = flags.pop('opt')
    save_pred_xlsx = flags.pop('save_pred_xlsx')
    skip_xlsx = flags.pop('skip_xlsx')
    measure_latency_flag = flags.pop('measure_latency_bs1')
    latency_warmup = flags.pop('latency_warmup')
    latency_max_samples = flags.pop('latency_max_samples')
    save_eff_json = flags.pop('save_eff_json')
    use_cfg_eval_dirs = flags.pop('use_cfg_eval_dirs')

    cfg.merge_dict(flags)
    cfg.merge_dict(opt)
    cfg = triple_base.prepare_cfg(cfg)

    if save_pred_xlsx is None:
        save_pred_xlsx = os.path.join(
            cfg.cfg['Global']['output_dir'],
            'preds_dump_text_triple_eff.xlsx',
        )
    if (not skip_xlsx) and save_pred_xlsx:
        save_pred_dir = os.path.dirname(save_pred_xlsx)
        if save_pred_dir:
            os.makedirs(save_pred_dir, exist_ok=True)
    if save_eff_json:
        save_eff_dir = os.path.dirname(save_eff_json)
        if save_eff_dir:
            os.makedirs(save_eff_dir, exist_ok=True)

    trainer = Trainer(cfg, mode='eval')

    if use_cfg_eval_dirs:
        data_dirs_list = []
        if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
            data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
        else:
            data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
            if data_dir_single:
                data_dirs_list = [[data_dir_single]]
        if not data_dirs_list:
            raise ValueError('No eval data dirs found in config while --use_cfg_eval_dirs=True.')
    else:
        data_dirs_list = [[
            r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_correct',
            r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_fakedv2',
        ]]

    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label_src', []),
        ('label_tgt', []),
        ('pred_text', []),
        ('pred_idsText', []),
        ('pred_ids', []),
        ('ids_valid', []),
        ('NED_text_src', []),
        ('NED_text_tgt', []),
        ('NED_idsText_src', []),
        ('NED_idsText_tgt', []),
    ])
    det_inputs_text = {'gts': [], 'preds': []}
    det_inputs_idsText = {'gts': [], 'preds': []}
    image_bytes = []

    total_num = 0
    total_src_text_true = 0
    total_tgt_text_true = 0
    total_src_text_neds = []
    total_tgt_text_neds = []

    total_src_idsText_true = 0
    total_tgt_idsText_true = 0
    total_src_idsText_neds = []
    total_tgt_idsText_neds = []

    legal_token_sum = 0
    total_token_sum = 0
    legal_seq_sum = 0
    total_seq_sum = 0

    total_eff_stats = None
    dataset_eff_payload = {}

    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            (
                src_pnacc_text, tgt_pnacc_text, src_ned_text_mean, tgt_ned_text_mean,
                src_pnacc_idsText, tgt_pnacc_idsText, src_ned_idsText_mean, tgt_ned_idsText_mean,
                num, legality, eff_summary,
            ) = dump_predictions(
                trainer,
                datadir,
                output_log,
                dataset_name,
                image_bytes,
                det_inputs_text,
                det_inputs_idsText,
                idc_arity=DEFAULT_IDC_ARITY,
                collect_artifacts=(not skip_xlsx),
            )

            print(f"{dataset_name}:\t text_src_acc: {100*src_pnacc_text:6g}, text_src_NED:{100*src_ned_text_mean:6g}, text_tgt_acc: {100*tgt_pnacc_text:6g}, text_tgt_NED:{100*tgt_ned_text_mean:6g}")
            print(f"{dataset_name}:\t idsText_src_acc: {100*src_pnacc_idsText:6g}, idsText_src_NED:{100*src_ned_idsText_mean:6g}, idsText_tgt_acc: {100*tgt_pnacc_idsText:6g}, idsText_tgt_NED:{100*tgt_ned_idsText_mean:6g}")

            latency_result = None
            if measure_latency_flag:
                latency_result = measure_latency_bs1(
                    trainer,
                    datadir,
                    dataset_name,
                    warmup=latency_warmup,
                    max_samples=latency_max_samples,
                )

            print_efficiency_summary(eff_summary, latency_result)
            print_efficiency_table_summary(dataset_name, eff_summary, latency_result)
            dataset_eff_payload[dataset_name] = {
                'efficiency': eff_summary,
                'latency_bs1': latency_result,
            }
            total_eff_stats = merge_efficiency_stats(total_eff_stats, eff_summary)

            total_num += num
            total_src_text_true += int(src_pnacc_text * num)
            total_tgt_text_true += int(tgt_pnacc_text * num)
            total_src_text_neds.extend([src_ned_text_mean] * num)
            total_tgt_text_neds.extend([tgt_ned_text_mean] * num)

            total_src_idsText_true += int(src_pnacc_idsText * num)
            total_tgt_idsText_true += int(tgt_pnacc_idsText * num)
            total_src_idsText_neds.extend([src_ned_idsText_mean] * num)
            total_tgt_idsText_neds.extend([tgt_ned_idsText_mean] * num)

            legal_token_sum += legality['legal_token']
            total_token_sum += legality['total_token']
            legal_seq_sum += legality['legal_seq']
            total_seq_sum += legality['total_seq']

    total_src_text_acc = (total_src_text_true / total_num) if total_num else 0.0
    total_tgt_text_acc = (total_tgt_text_true / total_num) if total_num else 0.0
    total_src_text_ned = float(np.mean(total_src_text_neds)) if total_src_text_neds else 0.0
    total_tgt_text_ned = float(np.mean(total_tgt_text_neds)) if total_tgt_text_neds else 0.0

    total_src_idsText_acc = (total_src_idsText_true / total_num) if total_num else 0.0
    total_tgt_idsText_acc = (total_tgt_idsText_true / total_num) if total_num else 0.0
    total_src_idsText_ned = float(np.mean(total_src_idsText_neds)) if total_src_idsText_neds else 0.0
    total_tgt_idsText_ned = float(np.mean(total_tgt_idsText_neds)) if total_tgt_idsText_neds else 0.0

    print(f"total (text):\t src_acc: {100*total_src_text_acc:6g}, src_NED:{100*total_src_text_ned:6g}, tgt_acc: {100*total_tgt_text_acc:6g}, tgt_NED:{100*total_tgt_text_ned:6g}")
    print(f"total (ids->text):\t src_acc: {100*total_src_idsText_acc:6g}, src_NED:{100*total_src_idsText_ned:6g}, tgt_acc: {100*total_tgt_idsText_acc:6g}, tgt_NED:{100*total_tgt_idsText_ned:6g}")

    if total_token_sum:
        token_legal_rate = legal_token_sum / total_token_sum
        print(f"IDS token legality: {token_legal_rate * 100:.2f}% ({legal_token_sum}/{total_token_sum})")
    if total_seq_sum:
        seq_legal_rate = legal_seq_sum / total_seq_sum
        print(f"IDS sequence legality: {seq_legal_rate * 100:.2f}% ({legal_seq_sum}/{total_seq_sum})")
    if (total_token_sum == 0) and (total_seq_sum == 0):
        print("[INFO] IDS legality skipped (no IDC tokens detected in predictions).")

    det_text = triple_base.calculate_cuo_metric_compact(det_inputs_text['gts'], det_inputs_text['preds'], X='X')
    det_idsText = triple_base.calculate_cuo_metric_compact(det_inputs_idsText['gts'], det_inputs_idsText['preds'], X='X')

    def fmt(x):
        return "N/A" if x is None else f"{x:.3f}"

    print("\nCuo detection metrics (text):")
    if det_text:
        print(f"N_sent={det_text['N_sent']} | clean={det_text['N_clean_sent']} | error={det_text['N_error_sent']}")
        print(f"Char_P={fmt(det_text['Char_P'])}%  Char_R={fmt(det_text['Char_R'])}%  Char_F1={fmt(det_text['Char_F1'])}%")
        print(f"Char_BalAcc={fmt(det_text['Char_BalAcc'])}%  Char_MCC={fmt(det_text['Char_MCC'])}")
        print(f"Sent_FA={fmt(det_text['Sent_FA'])}%  Sent_EM={fmt(det_text['Sent_EM'])}%")
    else:
        print("No sentences for text branch.")

    print("\nCuo detection metrics (ids->text):")
    if det_idsText:
        print(f"N_sent={det_idsText['N_sent']} | clean={det_idsText['N_clean_sent']} | error={det_idsText['N_error_sent']}")
        print(f"Char_P={fmt(det_idsText['Char_P'])}%  Char_R={fmt(det_idsText['Char_R'])}%  Char_F1={fmt(det_idsText['Char_F1'])}%")
        print(f"Char_BalAcc={fmt(det_idsText['Char_BalAcc'])}%  Char_MCC={fmt(det_idsText['Char_MCC'])}")
        print(f"Sent_FA={fmt(det_idsText['Sent_FA'])}%  Sent_EM={fmt(det_idsText['Sent_EM'])}%")
    else:
        print("No sentences for ids->text branch.")

    total_eff_summary = finalize_efficiency_stats(total_eff_stats) if total_eff_stats is not None else None
    total_latency_summary = aggregate_latency_results(dataset_eff_payload)
    if total_eff_summary is not None:
        print_efficiency_summary(total_eff_summary, total_latency_summary)
        print_efficiency_table_summary('total', total_eff_summary, total_latency_summary)

    if save_eff_json:
        with open(save_eff_json, 'w', encoding='utf-8') as f:
            json.dump({
                'datasets': dataset_eff_payload,
                'total': total_eff_summary,
                'total_latency_bs1': total_latency_summary,
            }, f, ensure_ascii=False, indent=2)
        print(f"[INFO] Efficiency summary saved to {save_eff_json}")

    if skip_xlsx:
        print("[INFO] XLSX dump skipped (--skip_xlsx=True).")
    else:
        try:
            maybe_save_xlsx(save_pred_xlsx, output_log, image_bytes)
        except Exception as e:
            print(f"[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.")


if __name__ == '__main__':
    main()
