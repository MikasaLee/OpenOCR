import io
import os
import sys
import torch
import math
import numpy as np
from collections import OrderedDict
from rapidfuzz.distance import Levenshtein
from tqdm import tqdm
from PIL import Image

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, '..')))

from tools.data import build_dataloader
from tools.engine.config import Config
from tools.engine.trainer import Trainer
from tools.utility import ArgsParser
from tools.utils.ids_syntax import validate_ids_prefix, DEFAULT_IDC_ARITY


# ============== Helpers ==============

def replace_punctuation(text: str) -> str:
    if text is None:
        return ''
    mapping = {
        '，': ',', '。': '.', '！': '!', '？': '?', '；': ';', '：': ':',
        '“': '"', '”': '"', '‘': "'", '’': "'",
    }
    for k, v in mapping.items():
        text = text.replace(k, v)
    return text


def to_png_bytes(img_array):
    if img_array is None:
        return None
    arr = np.array(img_array)
    if arr.ndim == 3 and arr.shape[0] in (1, 3):
        arr = np.transpose(arr, (1, 2, 0))
    arr = np.squeeze(arr)
    if arr.max() <= 1.0 + 1e-3:
        arr = arr * 255.0
    arr = np.clip(arr, 0, 255).astype(np.uint8)
    mode = 'L' if arr.ndim == 2 else 'RGB'
    img = Image.fromarray(arr, mode=mode)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    data = buf.getvalue()
    buf.close()
    return data


def split_src_tgt(gt_raw: str):
    if gt_raw is None:
        return '', ''
    parts = str(gt_raw).split('<unk>')  # 应该是\t，但ids字典中没有就转成了<unk>
    if len(parts) == 2:
        return parts[0], parts[1]
    return str(gt_raw), ''


def _safe_div(num, den):
    return None if den == 0 else (num / den)


def _safe_pct(num, den):
    v = _safe_div(num, den)
    return None if v is None else (v * 100.0)


def align_by_opcodes(gt: str, pred: str, gap_char=None):
    aligned = []
    for tag, i1, i2, j1, j2 in Levenshtein.opcodes(gt, pred):
        if tag in ("equal", "replace"):
            len_a = i2 - i1
            len_b = j2 - j1
            m = min(len_a, len_b)
            for k in range(m):
                gi = i1 + k
                pj = j1 + k
                aligned.append((gt[gi], pred[pj], gi))
            if len_a > m:
                for gi in range(i1 + m, i2):
                    aligned.append((gt[gi], gap_char, gi))
            if len_b > m:
                for pj in range(j1 + m, j2):
                    aligned.append((gap_char, pred[pj], None))
        elif tag == "delete":
            for gi in range(i1, i2):
                aligned.append((gt[gi], gap_char, gi))
        elif tag == "insert":
            for pj in range(j1, j2):
                aligned.append((gap_char, pred[pj], None))
        else:
            raise ValueError(f"Unknown opcode tag: {tag}")
    return aligned


def calculate_cuo_metric_compact(gt_sentences, pred_sentences, X='X'):
    """
    仅评 X（错字位点）的检测，输出精简但高信息量的指标：
    - N_clean_sent: GT 无 X 的句子数
    - N_error_sent: GT 有 X 的句子数
    - Char_P / Char_R / Char_F1:
        precision = TP/(TP+FP), recall = TP/(TP+FN)
    - Sent_FA: clean sentence 上预测出任意 X 的比例（误报率）
    - Sent_EM: error sentence 上 exact match（X位置集合完全一致且不能多报/插入X）的比例
    """
    n = min(len(gt_sentences), len(pred_sentences))
    if n == 0:
        return {}

    n_clean_sent = 0
    n_error_sent = 0
    sent_fa = 0   # false alarm on clean sentences
    sent_em = 0   # exact match on error sentences

    # char-level TP/FP/FN/TN
    # 修改：区分 对齐位置的FP(fp_align) 和 插入位置的FP(fp_ins)
    # 目的：计算 MCC/BalAcc 时只使用 GT 宇宙 (fp_align)，避免插入列带来混乱
    tp = 0
    fp_align = 0
    fp_ins = 0
    fn = 0
    tn = 0

    for gt, pred in zip(gt_sentences[:n], pred_sentences[:n]):
        aligned = align_by_opcodes(gt, pred, gap_char=None)

        gt_x_pos = set()
        pred_x_pos = set()
        ins_x_cnt = 0

        for gch, pch, gi in aligned:
            g_is_x = (gch == X)
            p_is_x = (pch == X)

            # 统计位置集合（以 GT 坐标为准）
            if gi is not None and g_is_x:
                gt_x_pos.add(gi)

            if p_is_x:
                if gi is None:
                    # pred 在 GT gap（插入列）上输出 X -> 一定是多报(FP) + exact match 必失败
                    ins_x_cnt += 1
                else:
                    pred_x_pos.add(gi)

            # char-level 计数
            if gi is None:
                # 插入列 (Insert)
                if p_is_x:
                    fp_ins += 1
                # else:
                #   GT gap, Pred NOT X. 
                #   用户建议：插入的非X不要算作 TN，以免大量插入非X“刷高”指标
                #   pass 
            else:
                # 对齐列 (GT universe)
                if g_is_x:
                    if p_is_x:
                        tp += 1
                    else:
                        fn += 1
                else:
                    # GT NOT X
                    if p_is_x:
                        fp_align += 1
                    else:
                        # GT NOT X, Pred NOT X -> TN for X
                        tn += 1

        gt_has_x = (len(gt_x_pos) > 0)
        pred_has_x = (len(pred_x_pos) > 0) or (ins_x_cnt > 0)

        if gt_has_x:
            n_error_sent += 1
            # exact match：X位置集合完全一致，且不能多报（含插入列X）
            if pred_has_x and (ins_x_cnt == 0) and (pred_x_pos == gt_x_pos):
                sent_em += 1
        else:
            n_clean_sent += 1
            # clean sentence 上只要预测出任意 X 就算误报
            if pred_has_x:
                sent_fa += 1

    # char-level metrics
    # P/R/F1: 保持原样，Precision 包含所有 FP (align + ins) 以反映模型实际输出
    fp_total = fp_align + fp_ins
    char_p = _safe_pct(tp, tp + fp_total)
    char_r = _safe_pct(tp, tp + fn)
    char_f1 = None
    if char_p is not None and char_r is not None and (char_p + char_r) > 0:
        char_f1 = 2 * char_p * char_r / (char_p + char_r)

    # Added: Balanced Accuracy & MCC
    # 修改：只在 GT 宇宙上计算 (使用 fp_align, 忽略 fp_ins 和 插入的TN)
    # Specificity = TN / (TN + FP_align)
    spec = _safe_div(tn, tn + fp_align)
    sens = _safe_div(tp, tp + fn) # Recall
    char_bal_acc = None
    if spec is not None and sens is not None:
        char_bal_acc = (spec + sens) / 2.0 * 100.0
    
    char_mcc = None
    # MCC 也只用 align 部分
    numerator = (tp * tn) - (fp_align * fn)
    denominator = (tp + fp_align) * (tp + fn) * (tn + fp_align) * (tn + fn)
    if denominator > 0:
        char_mcc = numerator / math.sqrt(denominator)

    # sentence-level metrics
    sent_fa_rate = _safe_pct(sent_fa, n_clean_sent)  # clean 误报率
    sent_em_rate = _safe_pct(sent_em, n_error_sent)  # error exact-match 率

    return {
        "N_sent": n,
        "N_clean_sent": n_clean_sent,
        "N_error_sent": n_error_sent,
        "Char_P": char_p,
        "Char_R": char_r,
        "Char_F1": char_f1,
        "Char_BalAcc": char_bal_acc,
        "Char_MCC": char_mcc,
        "Sent_FA": sent_fa_rate,
        "Sent_EM": sent_em_rate,
        # 下面三项只为方便定位（可不打印）
        "Char_TP": tp,
        "Char_FP": fp_total, # 打印总 FP
        "Char_FN": fn,
        "Char_TN": tn,
        "Char_FP_align": fp_align,
        "Char_FP_ins": fp_ins,
    }


def maybe_ids_tokens(tokens):
    if not tokens:
        return False
    idc_chars = {'⿰', '⿱', '⿲', '⿳', '⿴', '⿵', '⿶', '⿷', '⿸', '⿹', '⿺', '⿻'}
    return any(any(ch in idc_chars for ch in tok) for tok in tokens)


# ============== Args/CFG ==============

def parse_args():
    parser = ArgsParser()
    parser.add_argument('--save_pred_xlsx', type=str, default=None,
        help='Path to save prediction XLSX. Default: <output_dir>/preds_dump_text_triple.xlsx')
    args = parser.parse_args()
    return args


def prepare_cfg(cfg):
    if cfg.cfg['Global']['output_dir'][-1] == '/':
        cfg.cfg['Global']['output_dir'] = cfg.cfg['Global']['output_dir'][:-1]
    if cfg.cfg['Global']['pretrained_model'] is None:
        cfg.cfg['Global']['pretrained_model'] = cfg.cfg['Global']['output_dir'] + '/best.pth'
    cfg.cfg['Global']['use_amp'] = False
    cfg.cfg['PostProcess']['with_ratio'] = True
    cfg.cfg['Metric']['with_ratio'] = True
    cfg.cfg['Metric']['max_len'] = 30
    cfg.cfg['Metric']['max_len_ids'] = 200
    cfg.cfg['Metric']['max_single_char_ids_len'] = 50
    cfg.cfg['Global']['max_text_length'] = 30
    cfg.cfg['Global']['max_ids_length'] = 200
    cfg.cfg['Global']['max_single_char_ids_len'] = 50
    cfg.cfg['Metric']['max_ratio'] = 12
    
    # ---------------------------------------------------------
    # 自动注入 Eval 保存原图所需的配置 (SaveRawImageBytes, KeepKeys, collate_fn)
    # 这样就不需要在 yaml 中手动开关，避免训练时报错
    # ---------------------------------------------------------
    eval_transforms = cfg.cfg['Eval']['dataset']['transforms']

    # 1. 注入 SaveRawImageBytes (如果在 DecodeImagePIL 后)
    if not any('SaveRawImageBytes' in t for t in eval_transforms):
        insert_idx = 0
        for i, t in enumerate(eval_transforms):
            if 'DecodeImagePIL' in t:
                insert_idx = i + 1
                break
        eval_transforms.insert(insert_idx, {'SaveRawImageBytes': {'dst_key': 'image_raw', 'src_key': 'image'}})

    # 2. 设置 collate_fn
    if 'loader' not in cfg.cfg['Eval']:
        cfg.cfg['Eval']['loader'] = {}
    cfg.cfg['Eval']['loader']['collate_fn'] = 'RecWithRawCollator'

    # 3. 增强 KeepKeys
    # 尝试找到 KeepKeys 节点 (通常在 transforms 列表末尾)
    keep_keys_list = None
    for t in eval_transforms:
        if 'KeepKeys' in t:
            keep_keys_list = t['KeepKeys']['keep_keys']
            break
    
    # 兼容默认/旧逻辑（如果 loop 没找到，尝试直接取最后一个）
    if keep_keys_list is None and len(eval_transforms) > 0 and 'KeepKeys' in eval_transforms[-1]:
        keep_keys_list = eval_transforms[-1]['KeepKeys']['keep_keys']

    if keep_keys_list is not None:
        if 'real_ratio' not in keep_keys_list:
            keep_keys_list.append('real_ratio')
        if 'image_raw' not in keep_keys_list:
            keep_keys_list.append('image_raw')

    return cfg


# ============== Core Dump ==============

def dump_predictions(trainer, datadir, output_log, dataset_name, image_bytes, det_inputs_text, det_inputs_idsText, idc_arity=None):
    config_each = trainer.cfg.copy()
    if 'RatioDataSet' in config_each['Eval']['dataset']['name']:
        config_each['Eval']['dataset']['data_dir_list'] = [datadir]
    else:
        config_each['Eval']['dataset']['data_dir'] = datadir
    valid_dataloader = build_dataloader(config_each, 'Eval', trainer.logger)
    trainer.logger.info(f'{datadir} valid dataloader has {len(valid_dataloader)} iters')

    model = trainer.model
    device = trainer.device
    post_process = trainer.post_process_class
    model.eval()

    # DEBUG: Check config injection
    # print("\n[DEBUG] Transforms in dump_predictions:", config_each['Eval']['dataset']['transforms'])
    # print("[DEBUG] Collate fn:", config_each['Eval']['loader'].get('collate_fn'))
    # END DEBUG

    num = 0
    # text
    src_true_text = 0
    tgt_true_text = 0
    src_ned_text = []
    tgt_ned_text = []
    # text_from_ids
    src_true_idsText = 0
    tgt_true_idsText = 0
    src_ned_idsText = []
    tgt_ned_idsText = []
    # ids legality
    if idc_arity is None:
        idc_arity = DEFAULT_IDC_ARITY
    legal_token = 0
    total_token = 0
    legal_seq = 0
    total_seq = 0

    with torch.no_grad():
        pbar = tqdm(total=len(valid_dataloader), desc=f'eval {dataset_name}', position=0, leave=True)
        sample_offset = 0
        for batch_idx, batch in enumerate(valid_dataloader):
            # print(f"batch:{[(type(t), len(t) if hasattr(t, '__len__') else None) for t in batch]}")  # debug
            # if len(batch) > 6:
            #     print(f"batch[6] type first el: {type(batch[6][0]) if len(batch[6])>0 else 'empty'}")
            
            batch_tensor = [t.to(device) for t in batch[:6]]
            batch_numpy = [t.numpy() for t in batch[:6]]
            
            # Robustly find raw_images: look for a list of images (ndarray or bytes)
            raw_images = None
            if len(batch) > 6:
                # 遍历查找 (兼容插入了 real_ratio 等情况)
                for item in batch[6:]:
                    if isinstance(item, list) and len(item) > 0:
                        first_sample = item[0]
                        # Case 1: List of numpy arrays (H,W,C) or (H,W) - produced by SaveRawImageBytes
                        if isinstance(first_sample, np.ndarray) and first_sample.ndim >= 2:
                            raw_images = item
                            break
                        # Case 2: List of bytes - if specifically encoded
                        elif isinstance(first_sample, bytes):
                            raw_images = item
                            break
            
            # if raw_images is None:
            #    print("Warning: raw_images not found in batch!")

            preds = model(batch_tensor[0], data=batch_tensor[1:])
            post_result = post_process(preds, batch_numpy)
            # print(f"post_result:{post_result.keys()}")  # debug
            # Expect dict with keys: text, ids, text_from_ids
            gts = None
            outputs = None
            if isinstance(post_result, dict):
                outputs = post_result
                if 'label_text' in post_result:
                    gts = post_result['label_text']
            elif isinstance(post_result, list) and len(post_result) == 2 and isinstance(post_result[0], tuple):
                # Handle TextIDSLabelDecode dual-branch output: [(text_res, text_gt), (ids_res, ids_gt)]
                (text_res, text_gt), (ids_res, ids_gt) = post_result
                
                # Recover text from IDs manually since it's missing in list output
                idsText_res = []
                for item in ids_res:
                    ids_str = item[0] if isinstance(item, (tuple, list)) else item
                    conf = item[1] if isinstance(item, (tuple, list)) and len(item) > 1 else 1.0
                    
                    if hasattr(post_process, 'map_ids_to_text'):
                        idsText_res.append((post_process.map_ids_to_text(ids_str), conf))
                    else:
                        idsText_res.append(('', conf))
                
                outputs = {
                    'text': text_res,
                    'ids': ids_res,
                    'text_from_ids': idsText_res
                }
                gts = text_gt
            elif isinstance(post_result, tuple):
                # fallback: as in older scripts
                outputs, gts = post_result
            else:
                # single list fallback
                outputs = {'text': post_result}

            # Fetch lists per key
            text_list = outputs.get('text', [])
            ids_list = outputs.get('ids', [])
            idsText_list = outputs.get('text_from_ids', [])

            # Normalize to list of (str, prob)
            def norm_pairs(lst):
                out = []
                for item in lst:
                    if isinstance(item, (tuple, list)) and len(item) >= 1:
                        s = item[0]
                        p = item[1] if len(item) > 1 else 1.0
                        out.append((str(s), float(p)))
                    else:
                        out.append((str(item), 1.0))
                return out

            text_list = norm_pairs(text_list)
            ids_list = norm_pairs(ids_list)
            idsText_list = norm_pairs(idsText_list)

            # Obtain GTs (text) if provided by post_process
            if gts is None and isinstance(post_result, dict):
                # Some post decoders embed GTs via trainer; try get from dict
                gts = None
            # Fallback: derive GTs from dataloader batch_numpy decoded later via post_process; here gts stays None

            # Iterate samples by text_list length
            for i in range(len(text_list)):
                pred_text, _ = text_list[i]
                pred_ids, _ = ids_list[i] if i < len(ids_list) else ('', 1.0)
                pred_idsText, _ = idsText_list[i] if i < len(idsText_list) else ('', 1.0)

                gt_text = ''
                if gts is not None and i < len(gts):
                    gt_text = gts[i][0]
                # When GT not provided by post_process, try batch_numpy label
                if (gt_text is None or gt_text == '') and i < len(batch_numpy[1]):
                    try:
                        gt_text = batch_numpy[1][i].decode('utf-8') if isinstance(batch_numpy[1][i], bytes) else str(batch_numpy[1][i])
                    except Exception:
                        gt_text = str(batch_numpy[1][i])

                gt_src_txt, gt_tgt_txt = split_src_tgt(gt_text)
                # Normalize punctuation
                pred_text_norm = replace_punctuation(pred_text)
                pred_idsText_norm = replace_punctuation(pred_idsText)
                gt_src_norm = replace_punctuation(gt_src_txt)
                gt_tgt_norm = replace_punctuation(gt_tgt_txt)

                # print(f"pred_text_norm:{pred_text_norm}, pred_idsText_norm:{pred_idsText_norm}, gt_src_norm:{gt_src_norm}, gt_tgt_norm:{gt_tgt_norm}")  # debug
                # text metrics
                s_ned = 1 - Levenshtein.normalized_distance(pred_text_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                t_ned = 1 - Levenshtein.normalized_distance(pred_text_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_text.append(s_ned)
                tgt_ned_text.append(t_ned)
                if int(s_ned) == 1:
                    src_true_text += 1
                if int(t_ned) == 1:
                    tgt_true_text += 1

                # text_from_ids metrics
                s_ned_idsText = 1 - Levenshtein.normalized_distance(pred_idsText_norm, gt_src_norm) if gt_src_norm is not None else 0.0
                t_ned_idsText = 1 - Levenshtein.normalized_distance(pred_idsText_norm, gt_tgt_norm) if gt_tgt_norm is not None else 0.0
                src_ned_idsText.append(s_ned_idsText)
                tgt_ned_idsText.append(t_ned_idsText)
                if int(s_ned_idsText) == 1:
                    src_true_idsText += 1
                if int(t_ned_idsText) == 1:
                    tgt_true_idsText += 1

                num += 1

                # IDS legality (on pred_ids)
                tokens = [tok for tok in str(pred_ids).strip().split() if tok]
                
                # Check validity for all tokens (needed for Excel column)
                token_validity = []
                for tok in tokens:
                    ok, _, _ = validate_ids_prefix([ch for ch in tok if not ch.isspace()], idc_arity=idc_arity, require_closed=True)
                    token_validity.append(ok)
                
                is_sample_valid = all(token_validity)
                
                # Stats update (only if contains IDC chars)
                if maybe_ids_tokens(tokens):
                    for ok in token_validity:
                        total_token += 1
                        if ok:
                            legal_token += 1
                    if tokens:
                        total_seq += 1
                        if is_sample_valid:
                            legal_seq += 1

                # Typo detection inputs
                det_inputs_text['gts'].append(gt_src_norm)
                det_inputs_text['preds'].append(pred_text_norm)
                det_inputs_idsText['gts'].append(gt_src_norm)
                det_inputs_idsText['preds'].append(pred_idsText_norm)

                # Logging and image embedding
                img_name = f"{dataset_name}_{sample_offset + i}"
                output_log['img_name'].append(img_name)
                output_log['type'].append(dataset_name)
                output_log['label_src'].append(gt_src_norm)
                output_log['label_tgt'].append(gt_tgt_norm)
                output_log['pred_text'].append(pred_text_norm)
                output_log['pred_idsText'].append(pred_idsText_norm)
                output_log['pred_ids'].append(str(pred_ids))
                output_log['ids_valid'].append(is_sample_valid)
                output_log['NED_text_src'].append(float(s_ned))
                output_log['NED_text_tgt'].append(float(t_ned))
                output_log['NED_idsText_src'].append(float(s_ned_idsText))
                output_log['NED_idsText_tgt'].append(float(t_ned_idsText))

                try:
                    sample_img = raw_images[i] if raw_images is not None else None
                    image_bytes.append(to_png_bytes(sample_img))
                except Exception:
                    image_bytes.append(None)

            sample_offset += len(text_list)
            pbar.update(1)
        pbar.close()

    model.train()
    # Averages
    src_pnacc_text = src_true_text / num if num else 0.0
    tgt_pnacc_text = tgt_true_text / num if num else 0.0
    src_ned_text_mean = float(np.mean(src_ned_text)) if src_ned_text else 0.0
    tgt_ned_text_mean = float(np.mean(tgt_ned_text)) if tgt_ned_text else 0.0

    src_pnacc_idsText = src_true_idsText / num if num else 0.0
    tgt_pnacc_idsText = tgt_true_idsText / num if num else 0.0
    src_ned_idsText_mean = float(np.mean(src_ned_idsText)) if src_ned_idsText else 0.0
    tgt_ned_idsText_mean = float(np.mean(tgt_ned_idsText)) if tgt_ned_idsText else 0.0

    legality = {
        'legal_token': legal_token,
        'total_token': total_token,
        'legal_seq': legal_seq,
        'total_seq': total_seq,
    }

    return (
        src_pnacc_text, tgt_pnacc_text, src_ned_text_mean, tgt_ned_text_mean,
        src_pnacc_idsText, tgt_pnacc_idsText, src_ned_idsText_mean, tgt_ned_idsText_mean,
        num, legality,
    )


def main():
    FLAGS = parse_args()
    cfg = Config(FLAGS.config)
    FLAGS = vars(FLAGS)
    opt = FLAGS.pop('opt')
    cfg.merge_dict(FLAGS)
    cfg.merge_dict(opt)
    cfg = prepare_cfg(cfg)

    save_pred_xlsx = FLAGS.get('save_pred_xlsx')
    if save_pred_xlsx is None:
        save_pred_xlsx = os.path.join(cfg.cfg['Global']['output_dir'], 'preds_dump_text_triple.xlsx')
    os.makedirs(os.path.dirname(save_pred_xlsx), exist_ok=True)

    trainer = Trainer(cfg, mode='eval')

    data_dirs_list = []
    if cfg.cfg['Eval']['dataset'].get('data_dir_list', None):
        data_dirs_list = [cfg.cfg['Eval']['dataset']['data_dir_list']]
    else:
        data_dir_single = cfg.cfg['Eval']['dataset'].get('data_dir', None)
        if data_dir_single:
            data_dirs_list = [[data_dir_single]]

    # default two sets
    data_dirs_list = [[
        # r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/train_lmdb',
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_correct',
        r'/ipfs/lirunrui/lmdb_dataset/visual_c3_new_textline/test_lmdb/test_fakedv2',
    ]]

    output_log = OrderedDict([
        ('img_name', []),
        ('type', []),
        ('label_src', []),
        ('label_tgt', []),
        ('pred_text', []),
        ('pred_idsText', []),
        ('pred_ids', []),
        ('ids_valid', []),
        ('NED_text_src', []),
        ('NED_text_tgt', []),
        ('NED_idsText_src', []),
        ('NED_idsText_tgt', []),
    ])

    det_inputs_text = {'gts': [], 'preds': []}
    det_inputs_idsText = {'gts': [], 'preds': []}
    image_bytes = []

    # accumulators
    every_src_text_acc = []
    every_tgt_text_acc = []
    every_src_text_ned = []
    every_tgt_text_ned = []

    every_src_idsText_acc = []
    every_tgt_idsText_acc = []
    every_src_idsText_ned = []
    every_tgt_idsText_ned = []

    total_num = 0
    total_src_text_true = 0
    total_tgt_text_true = 0
    total_src_text_neds = []
    total_tgt_text_neds = []

    total_src_idsText_true = 0
    total_tgt_idsText_true = 0
    total_src_idsText_neds = []
    total_tgt_idsText_neds = []

    legal_token_sum = 0
    total_token_sum = 0
    legal_seq_sum = 0
    total_seq_sum = 0

    for data_dirs in data_dirs_list:
        for datadir in data_dirs:
            dataset_name = datadir[:-1].split('/')[-1] if datadir.endswith('/') else datadir.split('/')[-1]
            (
                src_pnacc_text, tgt_pnacc_text, src_ned_text_mean, tgt_ned_text_mean,
                src_pnacc_idsText, tgt_pnacc_idsText, src_ned_idsText_mean, tgt_ned_idsText_mean,
                num, legality,
            ) = dump_predictions(
                trainer, datadir, output_log, dataset_name, image_bytes,
                det_inputs_text, det_inputs_idsText, idc_arity=DEFAULT_IDC_ARITY,
            )

            print(f"{dataset_name}:\t text_src_acc: {100*src_pnacc_text:6g}, text_src_NED:{100*src_ned_text_mean:6g}, text_tgt_acc: {100*tgt_pnacc_text:6g}, text_tgt_NED:{100*tgt_ned_text_mean:6g}")
            print(f"{dataset_name}:\t idsText_src_acc: {100*src_pnacc_idsText:6g}, idsText_src_NED:{100*src_ned_idsText_mean:6g}, idsText_tgt_acc: {100*tgt_pnacc_idsText:6g}, idsText_tgt_NED:{100*tgt_ned_idsText_mean:6g}")

            every_src_text_acc.append(src_pnacc_text)
            every_tgt_text_acc.append(tgt_pnacc_text)
            every_src_text_ned.append(src_ned_text_mean)
            every_tgt_text_ned.append(tgt_ned_text_mean)

            every_src_idsText_acc.append(src_pnacc_idsText)
            every_tgt_idsText_acc.append(tgt_pnacc_idsText)
            every_src_idsText_ned.append(src_ned_idsText_mean)
            every_tgt_idsText_ned.append(tgt_ned_idsText_mean)

            total_num += num

            total_src_text_true += int(src_pnacc_text * num)
            total_tgt_text_true += int(tgt_pnacc_text * num)
            total_src_text_neds.extend([src_ned_text_mean] * num)
            total_tgt_text_neds.extend([tgt_ned_text_mean] * num)

            total_src_idsText_true += int(src_pnacc_idsText * num)
            total_tgt_idsText_true += int(tgt_pnacc_idsText * num)
            total_src_idsText_neds.extend([src_ned_idsText_mean] * num)
            total_tgt_idsText_neds.extend([tgt_ned_idsText_mean] * num)

            legal_token_sum += legality['legal_token']
            total_token_sum += legality['total_token']
            legal_seq_sum += legality['legal_seq']
            total_seq_sum += legality['total_seq']

    # Save XLSX
    try:
        import pandas as pd
        df = pd.DataFrame(output_log)
        df.to_excel(save_pred_xlsx, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.utils import get_column_letter
            wb = load_workbook(save_pred_xlsx)
            ws = wb.active
            img_col = ws.max_column + 1
            ws.cell(row=1, column=img_col, value='image')
            img_col_letter = get_column_letter(img_col)
            target_px = 160
            ws.column_dimensions[img_col_letter].width = max(ws.column_dimensions[img_col_letter].width or 0, target_px / 7.0)
            embedded = 0
            for r_idx, data in enumerate(image_bytes, start=2):
                if not data:
                    continue
                img_obj = OpenpyxlImage(io.BytesIO(data))
                try:
                    with Image.open(io.BytesIO(data)) as pil_img:
                        w, h = pil_img.size
                except Exception:
                    w, h = None, None
                img_obj.width = target_px
                if w and h and w > 0:
                    img_obj.height = h * (target_px / float(w))
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, img_obj.height * 0.75)
                img_obj.anchor = f"{img_col_letter}{r_idx}"
                ws.add_image(img_obj)
                embedded += 1
            print(f"[INFO] Embedded {embedded} images into Excel")
            wb.save(save_pred_xlsx)
        except Exception as embed_err:
            print(f"[WARN] Failed to embed images into XLSX ({embed_err}). Ensure openpyxl is installed.")

        # Aggregates & prints
        total_src_text_acc = (total_src_text_true / total_num) if total_num else 0.0
        total_tgt_text_acc = (total_tgt_text_true / total_num) if total_num else 0.0
        total_src_text_ned = float(np.mean(total_src_text_neds)) if total_src_text_neds else 0.0
        total_tgt_text_ned = float(np.mean(total_tgt_text_neds)) if total_tgt_text_neds else 0.0

        total_src_idsText_acc = (total_src_idsText_true / total_num) if total_num else 0.0
        total_tgt_idsText_acc = (total_tgt_idsText_true / total_num) if total_num else 0.0
        total_src_idsText_ned = float(np.mean(total_src_idsText_neds)) if total_src_idsText_neds else 0.0
        total_tgt_idsText_ned = float(np.mean(total_tgt_idsText_neds)) if total_tgt_idsText_neds else 0.0

        print(f"total (text):\t src_acc: {100*total_src_text_acc:6g}, src_NED:{100*total_src_text_ned:6g}, tgt_acc: {100*total_tgt_text_acc:6g}, tgt_NED:{100*total_tgt_text_ned:6g}")
        print(f"total (ids->text):\t src_acc: {100*total_src_idsText_acc:6g}, src_NED:{100*total_src_idsText_ned:6g}, tgt_acc: {100*total_tgt_idsText_acc:6g}, tgt_NED:{100*total_tgt_idsText_ned:6g}")
        if total_token_sum:
            token_legal_rate = legal_token_sum / total_token_sum
            print(f"IDS token legality: {token_legal_rate * 100:.2f}% ({legal_token_sum}/{total_token_sum})")
        if total_seq_sum:
            seq_legal_rate = legal_seq_sum / total_seq_sum
            print(f"IDS sequence legality: {seq_legal_rate * 100:.2f}% ({legal_seq_sum}/{total_seq_sum})")
        if (total_token_sum == 0) and (total_seq_sum == 0):
            print("[INFO] IDS legality skipped (no IDC tokens detected in predictions).")

        # Sent-level typo detection for text branch
        det_text = calculate_cuo_metric_compact(det_inputs_text['gts'], det_inputs_text['preds'], X='X')
        det_idsText = calculate_cuo_metric_compact(det_inputs_idsText['gts'], det_inputs_idsText['preds'], X='X')
        def fmt(x):
            return "N/A" if x is None else f"{x:.3f}"
        print("\nCuo detection metrics (text):")
        if det_text:
            print(f"N_sent={det_text['N_sent']} | clean={det_text['N_clean_sent']} | error={det_text['N_error_sent']}")
            print(f"Char_P={fmt(det_text['Char_P'])}%  Char_R={fmt(det_text['Char_R'])}%  Char_F1={fmt(det_text['Char_F1'])}%")
            print(f"Char_BalAcc={fmt(det_text['Char_BalAcc'])}%  Char_MCC={fmt(det_text['Char_MCC'])}")
            print(f"Sent_FA={fmt(det_text['Sent_FA'])}%  Sent_EM={fmt(det_text['Sent_EM'])}%")
        else:
            print("No sentences for text branch.")
        print("\nCuo detection metrics (ids->text):")
        if det_idsText:
            print(f"N_sent={det_idsText['N_sent']} | clean={det_idsText['N_clean_sent']} | error={det_idsText['N_error_sent']}")
            print(f"Char_P={fmt(det_idsText['Char_P'])}%  Char_R={fmt(det_idsText['Char_R'])}%  Char_F1={fmt(det_idsText['Char_F1'])}%")
            print(f"Char_BalAcc={fmt(det_idsText['Char_BalAcc'])}%  Char_MCC={fmt(det_idsText['Char_MCC'])}")
            print(f"Sent_FA={fmt(det_idsText['Sent_FA'])}%  Sent_EM={fmt(det_idsText['Sent_EM'])}%")
        else:
            print("No sentences for ids->text branch.")

        print(f"Predictions saved to {save_pred_xlsx}")
    except Exception as e:
        print(f"[WARN] Failed to save XLSX ({e}). Install pandas & openpyxl to enable XLSX export.")


if __name__ == '__main__':
    main()
